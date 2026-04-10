"""
run_consolidado.py
==================
Consolida la conciliación mensual de ambas cuentas Walmart MX (Atlas + Spring).

Uso:
    python run_consolidado.py --atlas Atlas_03-2026.csv --spring Spring_03-2026.csv

Opciones:
    --atlas   <csv>   CSV de pagos cuenta Atlas del Descanso (requerido)
    --spring  <csv>   CSV de pagos cuenta Spring (requerido)
    --output  -o      Nombre del Excel de salida (default: consolidado_<periodo>.xlsx)
    --skip-erp        Omite la consulta al ERP por completo

Ejemplo:
    python run_consolidado.py --atlas Atlas_03-2026.csv --spring Spring_03-2026.csv
    python run_consolidado.py --atlas Atlas_03-2026.csv --spring Spring_03-2026.csv --skip-erp

Dependencias:
    pip install pandas openpyxl requests python-dotenv
"""

import sys
import argparse
import importlib.util
from pathlib import Path
from datetime import datetime

_DIR = Path(__file__).parent
for _mod in ["conciliacion_walmart.py", "preparar_erp_walmart.py", "run_conciliacion.py"]:
    if not (_DIR / _mod).exists():
        print(f"ERROR: No se encontró {_mod} en {_DIR}")
        sys.exit(1)

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install pandas openpyxl requests python-dotenv")
    sys.exit(1)


def _load(name: str):
    spec = importlib.util.spec_from_file_location(name, _DIR / f"{name}.py")
    mod  = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod

_conc = _load("conciliacion_walmart")
_erp  = _load("preparar_erp_walmart")
_run  = _load("run_conciliacion")   # reutiliza ACCION_ERP, helpers de estilo y hoja_accion


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS LOCALES DE ESTILO
#  (equivalentes a los de run_conciliacion para uso en hoja_consolidado)
# ══════════════════════════════════════════════════════════════════════════════

def hfill(h):  return PatternFill("solid", start_color=h, fgColor=h)
def thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)
def mfmt():    return '#,##0.00;(#,##0.00);"-"'

C = dict(
    blue="2E75B6", blue2="0070C0", lgray="F2F2F2", white="FFFFFF",
    yellow="FFF2CC", secbg="D6E4F0",
)
F = {
    "title": Font(name="Arial", bold=True, size=14, color="FFFFFF"),
    "hdr":   Font(name="Arial", bold=True, size=10, color="FFFFFF"),
    "body":  Font(name="Arial", size=10),
    "bold":  Font(name="Arial", bold=True, size=10),
    "sec":   Font(name="Arial", bold=True, size=10, color="1F3864"),
    "ital":  Font(name="Arial", size=9,  italic=True, color="595959"),
}


# ══════════════════════════════════════════════════════════════════════════════
#  PROCESAR UNA CUENTA
# ══════════════════════════════════════════════════════════════════════════════

def procesar_cuenta(csv_path: str, cuenta: str, skip_erp: bool) -> tuple[pd.DataFrame, dict]:
    """Carga, enriquece y clasifica el CSV de una cuenta. Devuelve (df, sin_mapeo)."""
    print(f"\n  [{cuenta.upper()}] Leyendo CSV: {csv_path}")
    df = _conc.leer_csv(csv_path)
    print(f"  [{cuenta.upper()}] → {len(df):,} registros cargados")

    _, sin_mapeo = _conc.agrupar_por_categoria(df)
    df = _erp.preparar_para_erp(df, cuenta=cuenta)
    df = _run.enriquecer_df(df)

    if not skip_erp:
        print(f"  [{cuenta.upper()}] Consultando ERP...")
        df = _erp.enriquecer_con_erp(df, cuenta=cuenta)

    return df, sin_mapeo


# ══════════════════════════════════════════════════════════════════════════════
#  HOJA CONSOLIDADO
# ══════════════════════════════════════════════════════════════════════════════

def hoja_consolidado(wb, df_atlas: pd.DataFrame, df_spring: pd.DataFrame):
    """Hoja resumen con totales de ambas cuentas: acciones ERP + desglose por concepto."""
    ws = wb.create_sheet("📊 CONSOLIDADO", 0)

    N_COLS = 8
    for col, w in zip("ABCDEFGH", [30, 13, 20, 13, 20, 13, 20, 13]):
        ws.column_dimensions[col].width = w

    # Banner
    ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
    ws["A1"] = "CONSOLIDADO MENSUAL WALMART MX – ATLAS + SPRING"
    ws["A1"].font = F["title"]
    ws["A1"].fill = hfill(C["blue2"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Meta
    fecha = df_atlas["Payment Date"].dropna().iloc[0] if not df_atlas.empty else "N/D"
    for r, (lbl, val) in enumerate([
        ("Periodo de pago:", str(fecha)),
        ("Generado:",        datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Registros Atlas:", f"{len(df_atlas):,}"),
        ("Registros Spring:", f"{len(df_spring):,}"),
    ], 2):
        ws.cell(row=r, column=1, value=lbl).font = F["bold"]
        ws.cell(row=r, column=2, value=val).font  = F["body"]
        for ci in range(1, N_COLS + 1):
            ws.cell(row=r, column=ci).fill   = hfill(C["lgray"])
            ws.cell(row=r, column=ci).border = thin()
        ws.row_dimensions[r].height = 16

    # ── SECCIÓN 1: Resumen por Acción ERP ────────────────────────────────────
    r = 7
    ws.merge_cells(f"A{r}:{get_column_letter(N_COLS)}{r}")
    ws.cell(row=r, column=1, value="RESUMEN POR ACCIÓN ERP")
    ws.cell(row=r, column=1).font      = F["sec"]
    ws.cell(row=r, column=1).fill      = hfill(C["secbg"])
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(row=r, column=1).border    = thin()
    ws.row_dimensions[r].height = 18; r += 1

    for ci, h in enumerate(["Acción",
                             "Atlas Reg.", "Atlas Monto ($)",
                             "Spring Reg.", "Spring Monto ($)",
                             "Total Reg.", "Total Monto ($)", ""], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = F["hdr"]; c.fill = hfill(C["blue"])
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = thin()
    ws.row_dimensions[r].height = 24; r += 1

    tot_a_cnt = tot_a_monto = tot_s_cnt = tot_s_monto = 0

    for accion, label in _run.ETIQUETA_ACCION.items():
        bg, fg = _run.COLOR_ACCION[accion]
        a_mask  = df_atlas["Accion_ERP"]  == accion
        s_mask  = df_spring["Accion_ERP"] == accion
        a_cnt   = int(a_mask.sum())
        a_monto = df_atlas.loc[a_mask, "Amount"].sum()
        s_cnt   = int(s_mask.sum())
        s_monto = df_spring.loc[s_mask, "Amount"].sum()

        tot_a_cnt += a_cnt;  tot_a_monto += a_monto
        tot_s_cnt += s_cnt;  tot_s_monto += s_monto

        for ci, val in enumerate([label, a_cnt, a_monto, s_cnt, s_monto,
                                   a_cnt + s_cnt, a_monto + s_monto, ""], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = hfill(bg); c.border = thin()
            c.font = Font(name="Arial", size=10, bold=(ci == 1), color=fg)
            if ci in (3, 5, 7):
                c.number_format = mfmt(); c.alignment = Alignment(horizontal="right")
            elif ci in (2, 4, 6):
                c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 18; r += 1

    for ci, val in enumerate(["TOTAL",
                               tot_a_cnt, tot_a_monto,
                               tot_s_cnt, tot_s_monto,
                               tot_a_cnt + tot_s_cnt,
                               tot_a_monto + tot_s_monto, ""], 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.fill = hfill(C["yellow"]); c.border = thin(); c.font = F["bold"]
        if ci in (3, 5, 7):
            c.number_format = mfmt(); c.alignment = Alignment(horizontal="right")
        elif ci in (2, 4, 6):
            c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 20; r += 2

    # ── SECCIÓN 2: Desglose por Concepto ─────────────────────────────────────
    ws.merge_cells(f"A{r}:{get_column_letter(N_COLS)}{r}")
    ws.cell(row=r, column=1, value="DESGLOSE POR CONCEPTO")
    ws.cell(row=r, column=1).font      = F["sec"]
    ws.cell(row=r, column=1).fill      = hfill(C["secbg"])
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(row=r, column=1).border    = thin()
    ws.row_dimensions[r].height = 18; r += 1

    for ci, h in enumerate(["Concepto", "Tipo",
                             "Atlas ($)", "Atlas Reg.",
                             "Spring ($)", "Spring Reg.",
                             "Total ($)", "Total Reg."], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = F["hdr"]; c.fill = hfill(C["blue"])
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = thin()
    ws.row_dimensions[r].height = 24; r += 1

    # Combina conceptos de ambas cuentas
    def _agg(df):
        return (df.groupby(["Concept", "Income / Outcome"])["Amount"]
                  .agg(monto="sum", cnt="count")
                  .reset_index())

    combos_a = _agg(df_atlas).set_index(["Concept", "Income / Outcome"])
    combos_s = _agg(df_spring).set_index(["Concept", "Income / Outcome"])
    todos    = combos_a.index.union(combos_s.index)
    todos    = sorted(todos, key=lambda x: (x[1] == "Egreso", x[0]))  # Ingreso primero

    neto_a = neto_s = 0.0
    for concepto, tipo in todos:
        a_monto = combos_a.loc[(concepto, tipo), "monto"] if (concepto, tipo) in combos_a.index else 0.0
        a_cnt   = int(combos_a.loc[(concepto, tipo), "cnt"]) if (concepto, tipo) in combos_a.index else 0
        s_monto = combos_s.loc[(concepto, tipo), "monto"] if (concepto, tipo) in combos_s.index else 0.0
        s_cnt   = int(combos_s.loc[(concepto, tipo), "cnt"]) if (concepto, tipo) in combos_s.index else 0

        neto_a += a_monto; neto_s += s_monto
        fg   = "006100" if tipo == "Ingreso" else "C00000"
        fill = hfill(C["lgray"]) if r % 2 == 0 else hfill(C["white"])

        for ci, val in enumerate([concepto, tipo,
                                   a_monto, a_cnt,
                                   s_monto, s_cnt,
                                   a_monto + s_monto, a_cnt + s_cnt], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = fill; c.border = thin()
            c.font = Font(name="Arial", size=10, color=fg if ci <= 2 else "000000")
            if ci in (3, 5, 7):
                c.number_format = mfmt(); c.alignment = Alignment(horizontal="right")
            elif ci in (4, 6, 8):
                c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 16; r += 1

    # Fila neto total
    for ci, val in enumerate(["Neto Total", "",
                               neto_a, "",
                               neto_s, "",
                               neto_a + neto_s, ""], 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.fill = hfill(C["yellow"]); c.border = thin(); c.font = F["bold"]
        if ci in (3, 5, 7):
            c.number_format = mfmt(); c.alignment = Alignment(horizontal="right")
    ws.row_dimensions[r].height = 20


# ══════════════════════════════════════════════════════════════════════════════
#  HOJAS POR CUENTA
# ══════════════════════════════════════════════════════════════════════════════

# (accion_key, nombre hoja, título banner, color hex)
_HOJAS = [
    ("PAGAR_PROVEEDOR",  "A Pagar Proveedor",  "A PAGAR PROVEEDOR",    "006100"),
    ("NOTA_CREDITO",     "Notas de Crédito",   "NOTAS DE CRÉDITO",     "843C0C"),
    ("GASTO_PLATAFORMA", "Gastos Plataforma",  "GASTOS DE PLATAFORMA", "7F6000"),
    ("RETENCION_FISCAL", "Retenciones",        "RETENCIONES FISCALES", "1F3864"),
]


def hojas_combinadas(wb, df_atlas: pd.DataFrame, df_spring: pd.DataFrame,
                     sm_atlas: dict, sm_spring: dict):
    """Genera hojas consolidadas por tipo (Atlas + Spring), columna Cuenta al inicio."""
    df_atlas  = df_atlas.copy();  df_atlas["Cuenta"]  = "Atlas"
    df_spring = df_spring.copy(); df_spring["Cuenta"] = "Spring"
    df_all    = pd.concat([df_atlas, df_spring], ignore_index=True)

    # Hojas de acción (una por tipo, ambas cuentas)
    for accion_key, nombre_hoja, titulo_banner, color in _HOJAS:
        _run.hoja_accion(wb, df_all, accion_key, nombre_hoja, titulo_banner, color)

    # Pendientes combinados (Seller sin PO#)
    mask_pend = (
        df_all["FulfilledBy"].str.lower().str.contains("seller", na=False) &
        (df_all["ERP_ID"].isna() | df_all["ERP_ID"].eq(""))
    )
    df_pend = df_all[mask_pend].copy()
    if not df_pend.empty:
        ws = wb.create_sheet("Pendientes")
        cols_prio  = ["Cuenta", "Accion_Label", "Payment Date", "Orderline Number",
                      "Amount", "Concept", "Income / Outcome", "FulfilledBy",
                      "Invoice Number", "Return Number"]
        cols_extra = [c for c in df_pend.columns
                      if c not in cols_prio and c not in ("Accion_ERP",)]
        cols = [c for c in cols_prio if c in df_pend.columns] + cols_extra

        _run.set_col_widths(ws, cols)
        _run.banner(ws, "SELLER SIN PO# – Seguimiento manual", "C00000", len(cols), row=1)
        ws.cell(row=2, column=1, value="Registros:").font = _run.F["bold"]
        ws.cell(row=2, column=2, value=len(df_pend)).font = _run.F["body"]
        ws.cell(row=2, column=3, value="Monto:").font     = _run.F["bold"]
        ws.cell(row=2, column=4, value=df_pend["Amount"].sum()).font = _run.F["body"]
        ws.cell(row=2, column=4).number_format = mfmt()
        for ci in range(1, len(cols) + 1):
            ws.cell(row=2, column=ci).fill   = hfill(C["lgray"])
            ws.cell(row=2, column=ci).border = thin()
        ws.row_dimensions[2].height = 16
        ws.row_dimensions[3].height = 6
        _run.escribir_encabezados(ws, cols, row=4)
        _run.escribir_filas(ws, df_pend, cols, start_row=5)

    # ERP Conciliación combinada
    if "ERP_En_Sistema" in df_all.columns:
        _erp._hoja_erp_conciliacion(wb, df_all, nombre_hoja="ERP_Conciliacion")

    # Sin Mapeo combinado (fusiona ambas cuentas sumando montos por clave)
    sm_all = {}
    for d in [sm_atlas, sm_spring]:
        for k, v in d.items():
            sm_all[k] = sm_all.get(k, 0) + v
    if sm_all:
        ws = wb.create_sheet("Sin Mapeo")
        for col, w in zip("ABC", [30, 16, 16]):
            ws.column_dimensions[col].width = w
        for ci, h in enumerate(["Concepto CSV", "Tipo", "Total"], 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = F["hdr"]; c.fill = hfill(C["blue"])
            c.alignment = Alignment(horizontal="center"); c.border = thin()
        for ri, ((concepto, tipo), total) in enumerate(sm_all.items(), 2):
            alt = hfill(C["lgray"]) if ri % 2 == 0 else hfill(C["white"])
            for ci, val in enumerate([concepto, tipo, round(total, 2)], 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = F["body"]; c.fill = alt; c.border = thin()
                if ci == 3:
                    c.number_format = mfmt()
                    c.alignment = Alignment(horizontal="right")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Consolidado mensual de conciliación Walmart MX (Atlas + Spring)"
    )
    parser.add_argument("--atlas",    required=True,
                        help="CSV de pagos cuenta Atlas del Descanso")
    parser.add_argument("--spring",   required=True,
                        help="CSV de pagos cuenta Spring")
    parser.add_argument("--output", "-o", default="",
                        help="Nombre del Excel de salida (default: consolidado_<periodo>.xlsx)")
    parser.add_argument("--skip-erp", action="store_true",
                        help="Omite la consulta al ERP por completo")
    args = parser.parse_args()

    for path, nombre in [(args.atlas, "atlas"), (args.spring, "spring")]:
        if not Path(path).exists():
            print(f"\n  ERROR: No se encontró el CSV de {nombre}: {path}")
            sys.exit(1)

    print()
    print("  ╔══════════════════════════════════════════════════════════════╗")
    print("  ║    CONSOLIDADO MENSUAL WALMART MX  –  ATLAS + SPRING        ║")
    print("  ╚══════════════════════════════════════════════════════════════╝")

    # ── PASO 1: Procesar cada cuenta ─────────────────────────────────────────
    print("\n  ── PASO 1/3: Procesando cuentas ────────────────────────────────")
    df_atlas,  sm_atlas  = procesar_cuenta(args.atlas,  "atlas",  args.skip_erp)
    df_spring, sm_spring = procesar_cuenta(args.spring, "spring", args.skip_erp)

    # ── PASO 2: Resumen en consola ────────────────────────────────────────────
    print()
    print("  " + "═" * 64)
    print(f"  {'RESUMEN CONSOLIDADO':^64}")
    print("  " + "═" * 64)
    print(f"  {'Acción':<28} {'Atlas ($)':>16} {'Spring ($)':>16}")
    print("  " + "─" * 64)
    for accion, label in _run.ETIQUETA_ACCION.items():
        a = df_atlas.loc[df_atlas["Accion_ERP"]   == accion, "Amount"].sum()
        s = df_spring.loc[df_spring["Accion_ERP"] == accion, "Amount"].sum()
        print(f"  ▸ {label:<26} {a:>16,.2f} {s:>16,.2f}")
    print("  " + "─" * 64)
    print(f"  {'TOTAL':<28} "
          f"{df_atlas['Amount'].sum():>16,.2f} "
          f"{df_spring['Amount'].sum():>16,.2f}")
    print("  " + "═" * 64)

    # ── PASO 3: Generar Excel ─────────────────────────────────────────────────
    periodo = (df_atlas["Payment Date"].dropna().iloc[0]
               if not df_atlas.empty else datetime.now().strftime("%m-%Y"))
    output  = args.output or f"consolidado_{periodo}.xlsx"

    print(f"\n  ── PASO 3/3: Generando Excel ───────────────────────────────────")
    print(f"  → {output}")

    wb = Workbook()
    wb.remove(wb.active)

    hoja_consolidado(wb, df_atlas, df_spring)
    hojas_combinadas(wb, df_atlas, df_spring, sm_atlas, sm_spring)

    wb.save(output)
    print(f"  ✅ Listo: {output}")
    print()
    print("  Hojas generadas:")
    print("    📊 CONSOLIDADO        – Resumen total Atlas + Spring")
    print("    A Pagar Proveedor     – Órdenes liquidadas (col. Cuenta: Atlas/Spring)")
    print("    Notas de Crédito      – Devoluciones (col. Cuenta: Atlas/Spring)")
    print("    Gastos Plataforma     – Comisiones/WFS/SEM (col. Cuenta: Atlas/Spring)")
    print("    Retenciones           – IVA e ISR (col. Cuenta: Atlas/Spring)")
    if not args.skip_erp:
        print("    ERP_Conciliacion      – CSV vs ERP por orden ✅/⚠/❌ (col. Cuenta)")
    print()


if __name__ == "__main__":
    main()
