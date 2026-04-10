from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_, text
from typing import List, Dict, Any, Optional
from pydantic import BaseModel
import logging
import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import io
import subprocess
import uuid
import re
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import requests
import time
import random
import secrets

# Directorio raíz de los scripts de conciliación (un nivel arriba del portal)
CONCILIACION_DIR = Path(__file__).parent.parent

from database import get_db, engine, Base, SessionLocal
from models import PedidoMeli, CasoSeguimiento, HistoricoConciliacion, HistoricoOrden, Usuario
from config import get_settings
from datetime import datetime
from auth import hash_password, verify_password, create_access_token, get_current_user, require_admin

# Configuración de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuración
settings = get_settings()

# ========== CONFIGURACIÓN DE CORREO ==========
def _email_config():
    """Lee configuración de correo desde settings (cargados del .env)."""
    s = get_settings()
    return {
        'smtp_server': 'smtp.gmail.com',
        'smtp_port': 465,
        'sender':     s.email_sender,
        'password':   s.email_password,
        'recipients': [r.strip() for r in s.email_recipients.split(",") if r.strip()],
    }


# ========== CONFIGURACIÓN ML PAYMENTS API ==========
ML_TOKEN_URL = "http://localhost:8080/meli-token"
ML_TOKEN_API_KEY = get_settings().ml_token_api_key
ML_PAYMENTS_URL = "https://api.mercadopago.com/v1/payments/"
ML_SHIPMENTS_URL = "https://api.mercadolibre.com/shipments/"
ML_ORDERS_URL    = "https://api.mercadolibre.com/orders/"


# ========== MODELOS PYDANTIC PARA CORREO ==========
class PagoItem(BaseModel):
    REFERENCIA_ERP: str
    FECHA_LIBERACION: str = ""
    MONTO_VENTA: float = 0
    COMISION_MP: float = 0
    COSTO_ENVIO: float = 0
    NETO_RECIBIDO: float = 0
    METODO_PAGO: str = ""
    ESTATUS: str = ""
    TRANSACTION_TYPE: str = ""


class ResumenPagos(BaseModel):
    total_ordenes: int
    total_venta: float
    total_neto: float
    fecha: str


class EnviarCorreoRequest(BaseModel):
    pagos: List[PagoItem]
    resumen: ResumenPagos


# ========== MODELOS PYDANTIC PARA PAGOS CXC ==========
class PagoCxCItem(BaseModel):
    PEDIDO: str = ""
    FACTURA: str = ""
    CLIENTE: str = ""
    FORMA_COBRO: str = "MercadoLibre"
    TIPO: str = ""
    MONTO: float = 0
    REFERENCIA_ML: str = ""
    INGRESO_BRUTO: float = 0
    COMISION_MP: float = 0
    COSTO_ENVIO: float = 0
    FINANCIAMIENTO: float = 0
    IMPUESTOS: float = 0
    TOTAL_GASTOS: float = 0


class ResumenPagosCxC(BaseModel):
    total_pagos: int
    total_anticipos: int
    total_cobros_factura: int
    total_monto: float
    fecha: str


class EnviarCorreoCxCRequest(BaseModel):
    pagos: List[PagoCxCItem]
    resumen: ResumenPagosCxC


# ========== MODELOS PYDANTIC PARA PORTAL DE SEGUIMIENTO ==========
class HistorialItem(BaseModel):
    fecha: str
    tipo: str
    detalle: str


class DatosOrden(BaseModel):
    sourceIds: List[str] = []
    metodoPago: str = ""
    totalMovimientos: int = 0
    fechaLiberacion: str = ""


class CasoSeguimientoCreate(BaseModel):
    referencia: str
    tipoProblema: str
    estadoSeguimiento: str = "pendiente"
    montoNeto: float = 0
    montoBruto: float = 0
    costos: float = 0
    notas: str = ""
    enUltimoProceso: bool = True
    datosOrden: Dict[str, Any] = {}
    historial: List[Dict[str, Any]] = []


class CasoSeguimientoUpdate(BaseModel):
    tipoProblema: str = None
    estadoSeguimiento: str = None
    montoNeto: float = None
    montoBruto: float = None
    costos: float = None
    notas: str = None
    enUltimoProceso: bool = None
    datosOrden: Dict[str, Any] = None
    historial: List[Dict[str, Any]] = None


class SincronizarCasosRequest(BaseModel):
    casos: List[CasoSeguimientoCreate]


# ========== MODELOS PYDANTIC PARA HISTÓRICO ==========
class OrdenHistoricoItem(BaseModel):
    referenciaERP: str = ""
    estatus: str = ""
    datosOrden: Dict[str, Any] = {}


class GuardarConciliacionRequest(BaseModel):
    totalOrdenes: int = 0
    coincidencias: int = 0
    casosCreados: int = 0
    ordenesValidadas: int = 0
    montoEstadoCuenta: float = 0
    resumenPorEstatus: Dict[str, Any] = {}
    totalIngresosBruto: float = 0
    totalCostos: float = 0
    totalNetoReal: float = 0
    ordenes: List[OrdenHistoricoItem] = []
    notas: str = ""


# ========== MODELO PYDANTIC PAYMENTS BATCH ==========
class PaymentsBatchRequest(BaseModel):
    payment_ids: List[str]


# ========== MODELO PYDANTIC ORDERS BATCH ==========
class OrdersBatchRequest(BaseModel):
    order_ids: List[str]


# ========== MODELOS PYDANTIC AUTH / ADMIN ==========
class LoginRequest(BaseModel):
    username: str
    password: str


class CambiarPasswordRequest(BaseModel):
    password_actual: str
    password_nuevo: str


class UsuarioCreate(BaseModel):
    username: str
    nombre_completo: str
    password: str
    rol: str   # ADMIN | FINANZAS | LOGISTICA | VENTAS


class UsuarioUpdate(BaseModel):
    nombre_completo: Optional[str] = None
    rol: Optional[str] = None
    activo: Optional[bool] = None


# ========== HELPERS ML PAYMENTS API ==========

def obtener_token_ml() -> str:
    """Obtiene un access token válido de MercadoPago desde el endpoint admin."""
    resp = requests.get(
        ML_TOKEN_URL,
        headers={"x-api-key": ML_TOKEN_API_KEY},
        timeout=10,
        verify=False
    )
    resp.raise_for_status()
    data = resp.json()
    token = data.get("access_token")
    if not token:
        raise ValueError("La respuesta del token no contiene 'access_token'")
    return token


def consultar_payment(payment_id: str, token: str, max_retries: int = 5) -> Optional[Dict]:
    """Consulta un payment individual con retry/backoff ante rate limit y errores de red."""
    url = ML_PAYMENTS_URL + payment_id
    headers = {"Authorization": f"Bearer {token}"}

    for attempt in range(max_retries):
        try:
            resp = requests.get(url, headers=headers, timeout=15)

            if resp.status_code == 200:
                return resp.json()
            elif resp.status_code == 404:
                return None
            elif resp.status_code == 429:
                wait = (2 ** attempt) + random.uniform(0.5, 1.5)
                logger.warning(f"Rate limit en payment {payment_id}. Esperando {wait:.1f}s")
                time.sleep(wait)
            elif resp.status_code == 401:
                logger.error(f"Token inválido consultando payment {payment_id}")
                return None
            else:
                logger.warning(f"Error {resp.status_code} en payment {payment_id}")
                return None

        except requests.exceptions.ConnectionError as e:
            wait = (2 ** attempt) + random.uniform(1, 3)
            logger.warning(f"Conexión interrumpida (intento {attempt+1}/{max_retries}) payment {payment_id}. Esperando {wait:.1f}s")
            time.sleep(wait)
        except requests.exceptions.Timeout:
            wait = (2 ** attempt) + random.uniform(1, 3)
            logger.warning(f"Timeout (intento {attempt+1}/{max_retries}) payment {payment_id}. Esperando {wait:.1f}s")
            time.sleep(wait)
        except Exception as e:
            logger.error(f"Error inesperado consultando payment {payment_id}: {e}")
            return None

    logger.error(f"Falló después de {max_retries} intentos: payment {payment_id}")
    return None


def consultar_shipment(shipment_id: str, token: str, max_retries: int = 5) -> Optional[Dict]:
    """Consulta un shipment para obtener el order_id real."""
    url = ML_SHIPMENTS_URL + shipment_id
    headers = {"Authorization": f"Bearer {token}"}

    for attempt in range(max_retries):
        try:
            resp = requests.get(url, headers=headers, timeout=15)
            if resp.status_code == 200:
                return resp.json()
            elif resp.status_code == 404:
                logger.warning(f"Shipment {shipment_id} no encontrado")
                return None
            elif resp.status_code == 429:
                wait = (2 ** attempt) + random.uniform(0.5, 1.5)
                logger.warning(f"Rate limit en shipment {shipment_id}. Esperando {wait:.1f}s")
                time.sleep(wait)
            else:
                logger.warning(f"Error {resp.status_code} en shipment {shipment_id}")
                return None
        except requests.exceptions.ConnectionError as e:
            wait = (2 ** attempt) + random.uniform(1, 3)
            logger.warning(f"Conexión interrumpida (intento {attempt+1}/{max_retries}) shipment {shipment_id}. Esperando {wait:.1f}s")
            time.sleep(wait)
        except requests.exceptions.Timeout:
            wait = (2 ** attempt) + random.uniform(1, 3)
            logger.warning(f"Timeout (intento {attempt+1}/{max_retries}) shipment {shipment_id}. Esperando {wait:.1f}s")
            time.sleep(wait)
        except Exception as e:
            logger.error(f"Error inesperado consultando shipment {shipment_id}: {e}")
            return None

    logger.error(f"Falló después de {max_retries} intentos: shipment {shipment_id}")
    return None


def consultar_orden(order_id: str, token: str, max_retries: int = 5) -> Optional[Dict]:
    """Consulta una orden de ML para obtener todos sus payment IDs."""
    url = ML_ORDERS_URL + order_id
    headers = {"Authorization": f"Bearer {token}"}

    for attempt in range(max_retries):
        try:
            resp = requests.get(url, headers=headers, timeout=15)
            if resp.status_code == 200:
                return resp.json()
            elif resp.status_code == 404:
                return None
            elif resp.status_code == 429:
                wait = (2 ** attempt) + random.uniform(0.5, 1.5)
                logger.warning(f"Rate limit en orden {order_id}. Esperando {wait:.1f}s")
                time.sleep(wait)
            elif resp.status_code == 401:
                logger.error(f"Token inválido consultando orden {order_id}")
                return None
            else:
                logger.warning(f"Error {resp.status_code} en orden {order_id}")
                return None
        except requests.exceptions.ConnectionError:
            wait = (2 ** attempt) + random.uniform(1, 3)
            logger.warning(f"Conexión interrumpida (intento {attempt+1}/{max_retries}) orden {order_id}. Esperando {wait:.1f}s")
            time.sleep(wait)
        except requests.exceptions.Timeout:
            wait = (2 ** attempt) + random.uniform(1, 3)
            logger.warning(f"Timeout (intento {attempt+1}/{max_retries}) orden {order_id}. Esperando {wait:.1f}s")
            time.sleep(wait)
        except Exception as e:
            logger.error(f"Error inesperado consultando orden {order_id}: {e}")
            return None

    logger.error(f"Falló después de {max_retries} intentos: orden {order_id}")
    return None


def parsear_orden(data: Dict) -> Dict:
    """
    Extrae los payment IDs y montos de una orden de ML.
    Devuelve: { order_id, total_amount, payments: [{ id, amount, installments, status }] }
    """
    payments = data.get("payments") or []
    return {
        "order_id":     str(data.get("id", "")),
        "total_amount": data.get("total_amount") or data.get("paid_amount") or 0,
        "payments": [
            {
                "id":           str(p.get("id", "")),
                "amount":       p.get("transaction_amount") or 0,
                "installments": p.get("installments") or 1,
                "status":       p.get("status") or "",
                "status_detail":p.get("status_detail") or "",
            }
            for p in payments
        ]
    }


def parsear_payment(data: Dict) -> Dict:
    """
    Extrae los campos relevantes de la respuesta de la API de pagos.

    Desglose de charges_details:
      meli_fee      → comision       (cobrado al vendedor por ML)
      shp_*         → envio          (costo de envío cobrado al vendedor)
      *iva*         → iva            (IVA retenido por ML)
      *isr*         → isr            (ISR retenido por ML)
      financing*    → financiamiento (cuota MSI / meses sin intereses retenida al vendedor)

    Campos adicionales del estado de cuenta detectados en el frontend:
      retencion_ml       → dinero bloqueado por reclamo/disputa abierta
      ajuste_isr_iva     → cobros adicionales post-liberación (recobro retenciones)
      envio_comprador_sep→ envío del comprador depositado en liberación separada
    """
    charges = data.get("charges_details", [])
    comision = isr = iva = envio = financiamiento = 0.0

    for charge in charges:
        name = charge.get("name", "")
        amount   = charge.get("amounts", {}).get("original", 0) or 0
        refunded = charge.get("amounts", {}).get("refunded", 0) or 0
        net_amount = amount - refunded
        if name == "meli_fee":
            comision += net_amount
        elif "isr" in name:
            isr += net_amount
        elif "iva" in name:
            iva += net_amount
        elif name.startswith("shp_"):
            # Costo de envío cobrado al vendedor (shp_fulfillment, shp_me1, shp_me2, etc.)
            # Distinto a shipping_amount, que es lo que pagó el comprador
            envio += net_amount
        elif "financing" in name or "installment" in name:
            # Cuota de financiamiento por venta a meses sin intereses (MSI).
            # ML la retiene de la liquidación del vendedor y NO aparece como meli_fee.
            financiamiento += net_amount

    order = data.get("order") or {}
    transaction_details = data.get("transaction_details") or {}
    status = data.get("status") or ""

    # Para devoluciones: transaction_amount y net_received_amount reflejan la venta original,
    # no el estado actual. Como todo fue revertido, el ingreso neto real es 0.
    es_devolucion = status == "refunded"
    transaction_amount  = 0 if es_devolucion else (data.get("transaction_amount") or 0)
    net_received_amount = 0 if es_devolucion else (transaction_details.get("net_received_amount") or 0)
    shipping_amount     = 0 if es_devolucion else (data.get("shipping_amount") or 0)
    envio               = 0.0 if es_devolucion else envio
    financiamiento      = 0.0 if es_devolucion else financiamiento

    return {
        "payment_id":           str(data.get("id", "")),
        "order_id":             str(order.get("id", "")),
        "external_reference":   data.get("external_reference") or "",
        "status":               status,
        "status_detail":        data.get("status_detail") or "",
        "transaction_amount":   transaction_amount,
        "net_received_amount":  net_received_amount,
        "shipping_amount":      shipping_amount,
        "envio":                round(envio, 2),
        "comision":             round(comision, 2),
        "isr":                  round(isr, 2),
        "iva":                  round(iva, 2),
        "financiamiento":       round(financiamiento, 2),  # ← NUEVO: cargo MSI retenido por ML
        "money_release_status": data.get("money_release_status") or "",
        "money_release_date":   data.get("money_release_date") or "",
        "date_approved":        data.get("date_approved") or "",
        "payment_method_id":    data.get("payment_method_id") or "",
        "installments":         data.get("installments") or 1,
        "description":          data.get("description") or "",
    }


# Crear aplicación FastAPI
app = FastAPI(
    title="Conciliador MercadoLibre - API",
    description="API local para consultar pedidos desde PostgreSQL",
    version="1.0.0"
)

# Configurar CORS para permitir peticiones desde el HTML
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # En producción, especificar dominios permitidos
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Crear tablas si no existen (incluyendo casos_seguimiento y usuarios)
Base.metadata.create_all(bind=engine)


@app.on_event("startup")
def crear_admin_inicial():
    """Si no existe ningún usuario ADMIN, crear admin/Admin2024!"""
    db = SessionLocal()
    try:
        if not db.query(Usuario).filter_by(rol="ADMIN").first():
            admin = Usuario(
                username="admin",
                nombre_completo="Administrador",
                password_hash=hash_password("Admin2024!"),
                rol="ADMIN",
            )
            db.add(admin)
            db.commit()
            logger.info("Usuario admin inicial creado")
    finally:
        db.close()


# ========== ENDPOINTS AUTH ==========

@app.post("/auth/login")
def login(request: LoginRequest, db: Session = Depends(get_db)):
    """Autentica un usuario y devuelve un JWT."""
    user = db.query(Usuario).filter(Usuario.username == request.username).first()
    if not user or not verify_password(request.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    if not user.activo:
        raise HTTPException(status_code=401, detail="Usuario inactivo")

    user.last_login = datetime.utcnow()
    db.commit()

    token = create_access_token({"sub": user.username})
    return {"access_token": token, "token_type": "bearer", "user": user.to_dict()}


@app.get("/auth/me")
def auth_me(current_user: Usuario = Depends(get_current_user)):
    """Retorna datos del usuario autenticado."""
    return current_user.to_dict()


@app.post("/auth/cambiar-password")
def cambiar_password(
    request: CambiarPasswordRequest,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Cambia la contraseña del usuario autenticado."""
    if not verify_password(request.password_actual, current_user.password_hash):
        raise HTTPException(status_code=400, detail="Contraseña actual incorrecta")
    current_user.password_hash = hash_password(request.password_nuevo)
    db.commit()
    return {"success": True, "message": "Contraseña actualizada"}


# ========== ENDPOINTS ADMIN ==========
ROLES_VALIDOS = {"ADMIN", "FINANZAS", "LOGISTICA", "VENTAS"}


@app.get("/admin/usuarios")
def listar_usuarios(
    admin: Usuario = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Lista todos los usuarios del sistema."""
    usuarios = db.query(Usuario).order_by(Usuario.id).all()
    return [u.to_dict() for u in usuarios]


@app.post("/admin/usuarios")
def crear_usuario(
    data: UsuarioCreate,
    admin: Usuario = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Crea un nuevo usuario."""
    if data.rol not in ROLES_VALIDOS:
        raise HTTPException(status_code=400, detail=f"Rol inválido. Válidos: {ROLES_VALIDOS}")
    if db.query(Usuario).filter(Usuario.username == data.username).first():
        raise HTTPException(status_code=400, detail="El username ya existe")
    user = Usuario(
        username=data.username,
        nombre_completo=data.nombre_completo,
        password_hash=hash_password(data.password),
        rol=data.rol,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    logger.info(f"Usuario creado: {data.username} ({data.rol}) por admin {admin.username}")
    return user.to_dict()


@app.put("/admin/usuarios/{user_id}")
def editar_usuario(
    user_id: int,
    data: UsuarioUpdate,
    admin: Usuario = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Edita nombre, rol o estado activo de un usuario."""
    try:
        user = db.query(Usuario).filter(Usuario.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        if data.nombre_completo is not None:
            user.nombre_completo = data.nombre_completo
        if data.rol is not None:
            if data.rol not in ROLES_VALIDOS:
                raise HTTPException(status_code=400, detail=f"Rol inválido. Válidos: {ROLES_VALIDOS}")
            user.rol = data.rol
        if data.activo is not None:
            user.activo = data.activo
        db.commit()
        db.refresh(user)
        return user.to_dict()
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error editando usuario {user_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.post("/admin/usuarios/{user_id}/reset-password")
def reset_password(
    user_id: int,
    admin: Usuario = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Genera y asigna una contraseña temporal al usuario."""
    user = db.query(Usuario).filter(Usuario.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    temp_password = secrets.token_urlsafe(10)
    user.password_hash = hash_password(temp_password)
    db.commit()
    logger.info(f"Password reseteado para usuario {user.username} por admin {admin.username}")
    return {"success": True, "tempPassword": temp_password, "username": user.username}


@app.delete("/admin/usuarios/{user_id}")
def desactivar_usuario(
    user_id: int,
    admin: Usuario = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Desactiva un usuario (soft delete)."""
    user = db.query(Usuario).filter(Usuario.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if user.id == admin.id:
        raise HTTPException(status_code=400, detail="No puedes desactivarte a ti mismo")
    user.activo = False
    db.commit()
    logger.info(f"Usuario desactivado: {user.username} por admin {admin.username}")
    return {"success": True, "message": f"Usuario {user.username} desactivado"}


@app.get("/")
def read_root():
    """Endpoint raíz para verificar que la API está funcionando."""
    return {
        "message": "API Conciliador MercadoLibre",
        "status": "running",
        "version": "1.0.0"
    }


@app.post("/payments/batch")
def payments_batch(request: PaymentsBatchRequest, current_user: Usuario = Depends(get_current_user)):
    """
    Enriquece una lista de payment IDs consultando la API de MercadoPago.
    Retorna un dict { payment_id: datos_parseados | null }.
    """
    if not request.payment_ids:
        return {"results": {}}

    try:
        token = obtener_token_ml()
    except Exception as e:
        logger.error(f"No se pudo obtener token ML: {e}")
        raise HTTPException(status_code=502, detail=f"Error obteniendo token MercadoPago: {e}")

    results = {}
    total = len(request.payment_ids)

    for i, pid in enumerate(request.payment_ids):
        data = consultar_payment(pid, token)
        parsed = parsear_payment(data) if data else None

        # Si es un pago de envío, el order.id es en realidad un shipment_id.
        # Consultamos el shipment para obtener el order_id real.
        if parsed and parsed.get("description") == "marketplace_shipment":
            shipment_id = str(data.get("order", {}).get("id", ""))
            if shipment_id:
                shipment_data = consultar_shipment(shipment_id, token)
                if shipment_data and shipment_data.get("order_id"):
                    real_order_id = str(shipment_data["order_id"])
                    parsed["order_id"] = real_order_id
                    parsed["external_reference"] = real_order_id
                    parsed["shipment_id"] = shipment_id
                else:
                    parsed["shipment_id"] = shipment_id

        results[pid] = parsed

        # Pausa variable anti rate-limit (lotes pequeños de 10, podemos ir más rápido)
        time.sleep(random.uniform(0.2, 0.5))

        # Pausa larga cada 100 consultas
        if (i + 1) % 100 == 0:
            logger.info(f"Pausa larga tras {i+1}/{total} consultas...")
            time.sleep(random.uniform(5, 8))

    logger.info(f"Batch completado: {sum(1 for v in results.values() if v)} de {total} encontrados")
    return {"results": results}


@app.post("/orders/batch")
def orders_batch(request: OrdersBatchRequest, current_user: Usuario = Depends(get_current_user)):
    """
    Consulta una lista de order_ids en la API de MercadoLibre.
    Retorna un dict { order_id: { order_id, total_amount, payments: [...] } | null }.
    Usado para detectar payments pendientes de liberar (ej. MSI) que no aparecen
    en el estado de cuenta del período actual.
    """
    if not request.order_ids:
        return {"results": {}}

    try:
        token = obtener_token_ml()
    except Exception as e:
        logger.error(f"No se pudo obtener token ML: {e}")
        raise HTTPException(status_code=502, detail=f"Error obteniendo token MercadoPago: {e}")

    results = {}
    total = len(request.order_ids)

    for i, oid in enumerate(request.order_ids):
        data = consultar_orden(oid, token)
        results[oid] = parsear_orden(data) if data else None

        time.sleep(random.uniform(0.2, 0.5))

        if (i + 1) % 100 == 0:
            logger.info(f"Orders batch: pausa larga tras {i+1}/{total} consultas...")
            time.sleep(random.uniform(5, 8))

    logger.info(f"Orders batch completado: {sum(1 for v in results.values() if v)} de {total} encontrados")
    return {"results": results}


@app.get("/ventas/venta-id/{venta_id}")
def get_venta_by_id(venta_id: str, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """
    Consulta 2: Búsqueda por EXTERNAL_REFERENCE (venta_id).

    Busca en:
    - referencia con patrón {venta_id}-%
    - adicional5 como valor único o en lista separada por comas

    Args:
        venta_id: ID de venta de Mercado Libre (EXTERNAL_REFERENCE)
        db: Sesión de base de datos

    Returns:
        Lista de pedidos encontrados
    """
    try:
        logger.info(f"Buscando venta_id: {venta_id}")

        # Parámetros para la búsqueda
        base_pattern = f"{venta_id}-%"
        adicional_exact = venta_id
        adicional_middle = f"%,{venta_id},%"
        adicional_start = f"{venta_id},%"
        adicional_end = f"%,{venta_id}"

        # Construir query
        query = db.query(PedidoMeli).filter(
            or_(
                # Coincidencia en referencia
                PedidoMeli.referencia.like(base_pattern),
                # Coincidencia en adicional5 (solo si referencia NO coincide)
                and_(
                    ~PedidoMeli.referencia.like(base_pattern),
                    or_(
                        PedidoMeli.adicional5 == adicional_exact,
                        PedidoMeli.adicional5.like(adicional_middle),
                        PedidoMeli.adicional5.like(adicional_start),
                        PedidoMeli.adicional5.like(adicional_end)
                    )
                )
            )
        )

        pedidos = query.all()

        # Deduplicar por (referencia, movid) — la tabla puede tener filas duplicadas
        vistos = set()
        pedidos_unicos = []
        for p in pedidos:
            key = (p.referencia, p.movid)
            if key not in vistos:
                vistos.add(key)
                pedidos_unicos.append(p)
        pedidos = pedidos_unicos

        if not pedidos:
            logger.info(f"No se encontraron pedidos para venta_id: {venta_id}")
            raise HTTPException(status_code=404, detail="No se encontraron pedidos")

        logger.info(f"Encontrados {len(pedidos)} pedidos para venta_id: {venta_id}")

        # Convertir a diccionario
        result = [pedido.to_dict() for pedido in pedidos]
        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en get_venta_by_id: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")


@app.get("/ventas/atencion/{atencion_id}")
def get_venta_by_atencion(atencion_id: str, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """
    Consulta 1: Búsqueda directa por campo atencion.

    Busca pedidos donde el campo atencion coincide exactamente con el identificador.
    Puede contener EXTERNAL_REFERENCE o SOURCE_ID.

    Args:
        atencion_id: Identificador en campo atencion (puede ser EXTERNAL_REFERENCE o SOURCE_ID)
        db: Sesión de base de datos

    Returns:
        Lista de pedidos encontrados
    """
    try:
        logger.info(f"Buscando atencion_id: {atencion_id}")

        pedidos = db.query(PedidoMeli).filter(
            PedidoMeli.atencion == atencion_id
        ).all()

        # Deduplicar por (referencia, movid) — la tabla puede tener filas duplicadas
        vistos = set()
        pedidos_unicos = []
        for p in pedidos:
            key = (p.referencia, p.movid)
            if key not in vistos:
                vistos.add(key)
                pedidos_unicos.append(p)
        pedidos = pedidos_unicos

        if not pedidos:
            logger.info(f"No se encontraron pedidos para atencion_id: {atencion_id}")
            raise HTTPException(status_code=404, detail="No se encontraron pedidos")

        logger.info(f"Encontrados {len(pedidos)} pedidos para atencion_id: {atencion_id}")

        # Convertir a diccionario
        result = [pedido.to_dict() for pedido in pedidos]
        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en get_venta_by_atencion: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")


@app.get("/health")
def health_check(db: Session = Depends(get_db)):
    """
    Endpoint de salud para verificar la conexión a la base de datos.
    """
    try:
        # Intentar una query simple
        db.execute(text("SELECT 1"))
        return {
            "status": "healthy",
            "database": "connected"
        }
    except Exception as e:
        logger.error(f"Error en health check: {str(e)}")
        raise HTTPException(
            status_code=503,
            detail=f"Base de datos no disponible: {str(e)}"
        )


# ========== FUNCIONES DE CORREO ==========
def crear_excel_pagos(pagos: List[Dict[str, Any]], resumen: Dict[str, Any]) -> bytes:
    """Genera un archivo Excel con los pagos"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos MercadoLibre"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    money_format = '#,##0.00'

    # Título
    ws.merge_cells('A1:I1')
    ws['A1'] = f"PAGOS MERCADOLIBRE - {resumen['fecha']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Resumen
    ws['A3'] = "Total Órdenes:"
    ws['B3'] = resumen['total_ordenes']
    ws['C3'] = "Total Venta:"
    ws['D3'] = resumen['total_venta']
    ws['D3'].number_format = money_format
    ws['E3'] = "Total Neto:"
    ws['F3'] = resumen['total_neto']
    ws['F3'].number_format = money_format

    # Headers de la tabla
    headers = ['Referencia ERP', 'Fecha', 'Monto Venta', 'Comisión MP', 'Costo Envío',
               'Neto Recibido', 'Método Pago', 'Estatus', 'Tipo Transacción']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Datos
    for row_idx, pago in enumerate(pagos, 6):
        ws.cell(row=row_idx, column=1, value=pago.get('REFERENCIA_ERP', '')).border = border

        fecha = pago.get('FECHA_LIBERACION', '')
        if fecha and len(fecha) >= 10:
            fecha = fecha[:10]
        ws.cell(row=row_idx, column=2, value=fecha).border = border

        cell_venta = ws.cell(row=row_idx, column=3, value=pago.get('MONTO_VENTA', 0))
        cell_venta.number_format = money_format
        cell_venta.border = border

        cell_comision = ws.cell(row=row_idx, column=4, value=pago.get('COMISION_MP', 0))
        cell_comision.number_format = money_format
        cell_comision.border = border

        cell_envio = ws.cell(row=row_idx, column=5, value=pago.get('COSTO_ENVIO', 0))
        cell_envio.number_format = money_format
        cell_envio.border = border

        cell_neto = ws.cell(row=row_idx, column=6, value=pago.get('NETO_RECIBIDO', 0))
        cell_neto.number_format = money_format
        cell_neto.border = border
        cell_neto.font = Font(bold=True)

        ws.cell(row=row_idx, column=7, value=pago.get('METODO_PAGO', '')).border = border
        ws.cell(row=row_idx, column=8, value=pago.get('ESTATUS', '')).border = border
        ws.cell(row=row_idx, column=9, value=pago.get('TRANSACTION_TYPE', '')).border = border

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18

    # Guardar a bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def enviar_correo_con_adjunto(pagos: List[Dict], resumen: Dict) -> bool:
    """Envía el correo con el archivo Excel adjunto"""

    # Crear el Excel
    excel_bytes = crear_excel_pagos(pagos, resumen)

    # Crear mensaje
    msg = MIMEMultipart()
    msg['From'] = _email_config()['sender']
    msg['To'] = ', '.join(_email_config()['recipients'])
    msg['Subject'] = f"Pagos MercadoLibre - {resumen['fecha']} - {resumen['total_ordenes']} órdenes"

    # Cuerpo del correo
    body = f"""
Buen día,

Se adjunta el archivo de pagos de MercadoLibre para su procesamiento.

RESUMEN:
- Fecha: {resumen['fecha']}
- Total de órdenes: {resumen['total_ordenes']}
- Monto total de ventas: ${resumen['total_venta']:,.2f}
- Neto a recibir: ${resumen['total_neto']:,.2f}

Por favor procesar los pagos correspondientes en el ERP.

Saludos,
Sistema de Conciliación MercadoLibre
    """
    msg.attach(MIMEText(body, 'plain'))

    # Adjuntar Excel
    attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    attachment.set_payload(excel_bytes)
    encoders.encode_base64(attachment)

    filename = f"Pagos_MercadoLibre_{resumen['fecha']}.xlsx"
    attachment.add_header('Content-Disposition', f'attachment; filename="{filename}"')
    msg.attach(attachment)

    # Enviar correo
    context = ssl.create_default_context()

    with smtplib.SMTP_SSL(
        _email_config()['smtp_server'],
        _email_config()['smtp_port'],
        context=context
    ) as server:
        server.login(_email_config()['sender'], _email_config()['password'])
        server.sendmail(
            _email_config()['sender'],
            _email_config()['recipients'],
            msg.as_string()
        )

    return True


@app.post("/enviar-correo-pagos")
async def enviar_correo_pagos(request: EnviarCorreoRequest, current_user: Usuario = Depends(get_current_user)):
    """Endpoint para enviar correo con los pagos de MercadoLibre"""
    try:
        logger.info(f"Enviando correo con {len(request.pagos)} pagos")

        pagos_dict = [p.dict() for p in request.pagos]
        resumen_dict = request.resumen.dict()

        enviar_correo_con_adjunto(pagos_dict, resumen_dict)

        logger.info(f"Correo enviado exitosamente a {_email_config()['recipients']}")

        return {
            "success": True,
            "message": f"Correo enviado exitosamente a {len(_email_config()['recipients'])} destinatarios con {len(pagos_dict)} pagos"
        }
    except smtplib.SMTPAuthenticationError:
        logger.error("Error de autenticación SMTP")
        raise HTTPException(status_code=500, detail="Error de autenticación con el servidor de correo")
    except smtplib.SMTPException as e:
        logger.error(f"Error SMTP: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error SMTP: {str(e)}")
    except Exception as e:
        logger.error(f"Error enviando correo: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error enviando correo: {str(e)}")


# ========== FUNCIONES DE CORREO PARA PAGOS CXC ==========
def crear_excel_pagos_cxc(pagos: List[Dict[str, Any]], resumen: Dict[str, Any]) -> bytes:
    """Genera un archivo Excel con los pagos CxC (Anticipos y Cobros a Factura)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos CxC"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill_anticipo = PatternFill(start_color="92400E", end_color="92400E", fill_type="solid")
    header_fill_cobro = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    money_format = '#,##0.00'

    # Título
    ws.merge_cells('A1:M1')
    ws['A1'] = f"PAGOS CXC MERCADOLIBRE - {resumen['fecha']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Resumen
    ws['A3'] = "Total Pagos:"
    ws['B3'] = resumen['total_pagos']
    ws['C3'] = "Anticipos:"
    ws['D3'] = resumen['total_anticipos']
    ws['E3'] = "Cobros Factura:"
    ws['F3'] = resumen['total_cobros_factura']

    ws['A4'] = "Monto Total:"
    ws['B4'] = resumen['total_monto']
    ws['B4'].number_format = money_format

    # Headers de la tabla
    headers = ['Pedido', 'Factura', 'Cliente', 'Forma de cobro', 'Tipo', 'Monto',
               'Ingreso Bruto', 'Comision MP', 'Costo Envio', 'Financiamiento', 'Impuestos',
               'Total Gastos', 'Referencia ML']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill_cobro
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Datos
    for row_idx, pago in enumerate(pagos, 7):
        ws.cell(row=row_idx, column=1, value=pago.get('PEDIDO', '')).border = border
        ws.cell(row=row_idx, column=2, value=pago.get('FACTURA', '')).border = border
        ws.cell(row=row_idx, column=3, value=pago.get('CLIENTE', '')).border = border
        ws.cell(row=row_idx, column=4, value=pago.get('FORMA_COBRO', 'MercadoLibre')).border = border
        ws.cell(row=row_idx, column=5, value=pago.get('TIPO', '')).border = border

        cell_monto = ws.cell(row=row_idx, column=6, value=pago.get('MONTO', 0))
        cell_monto.number_format = money_format
        cell_monto.border = border
        cell_monto.font = Font(bold=True)

        cell_bruto = ws.cell(row=row_idx, column=7, value=pago.get('INGRESO_BRUTO', 0))
        cell_bruto.number_format = money_format
        cell_bruto.border = border

        cell_com = ws.cell(row=row_idx, column=8, value=pago.get('COMISION_MP', 0))
        cell_com.number_format = money_format
        cell_com.border = border

        cell_env = ws.cell(row=row_idx, column=9, value=pago.get('COSTO_ENVIO', 0))
        cell_env.number_format = money_format
        cell_env.border = border

        cell_fin = ws.cell(row=row_idx, column=10, value=pago.get('FINANCIAMIENTO', 0))
        cell_fin.number_format = money_format
        cell_fin.border = border

        cell_imp = ws.cell(row=row_idx, column=11, value=pago.get('IMPUESTOS', 0))
        cell_imp.number_format = money_format
        cell_imp.border = border

        cell_tg = ws.cell(row=row_idx, column=12, value=pago.get('TOTAL_GASTOS', 0))
        cell_tg.number_format = money_format
        cell_tg.border = border
        cell_tg.font = Font(bold=True)

        ws.cell(row=row_idx, column=13, value=pago.get('REFERENCIA_ML', '')).border = border

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 20

    # Guardar a bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def enviar_correo_pagos_cxc_adjunto(pagos: List[Dict], resumen: Dict) -> bool:
    """Envía el correo con el archivo Excel de pagos CxC adjunto"""

    # Crear el Excel
    excel_bytes = crear_excel_pagos_cxc(pagos, resumen)

    # Crear mensaje
    msg = MIMEMultipart()
    msg['From'] = _email_config()['sender']
    msg['To'] = ', '.join(_email_config()['recipients'])
    msg['Subject'] = f"Pagos CxC MercadoLibre - {resumen['fecha']} - {resumen['total_anticipos']} anticipos, {resumen['total_cobros_factura']} cobros"

    # Cuerpo del correo
    body = f"""
Buen día,

Se adjunta el archivo de PAGOS CXC de MercadoLibre para su procesamiento.

RESUMEN:
- Fecha: {resumen['fecha']}
- Total de pagos: {resumen['total_pagos']}
- Anticipos CxC (Solo Pedido Pendiente): {resumen['total_anticipos']}
- Cobros a Factura (Factura Concluida): {resumen['total_cobros_factura']}
- Monto total: ${resumen['total_monto']:,.2f}

INSTRUCCIONES:
- ANTICIPOS: Aplicar anticipo de CxC al Pedido indicado.
- COBROS A FACTURA: Aplicar cobro a la Factura indicada.

Por favor procesar los pagos correspondientes en el ERP.

Saludos,
Sistema de Conciliación MercadoLibre
    """
    msg.attach(MIMEText(body, 'plain'))

    # Adjuntar Excel
    attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    attachment.set_payload(excel_bytes)
    encoders.encode_base64(attachment)

    filename = f"Pagos_CxC_MercadoLibre_{resumen['fecha']}.xlsx"
    attachment.add_header('Content-Disposition', f'attachment; filename="{filename}"')
    msg.attach(attachment)

    # Enviar correo
    context = ssl.create_default_context()

    with smtplib.SMTP_SSL(
        _email_config()['smtp_server'],
        _email_config()['smtp_port'],
        context=context
    ) as server:
        server.login(_email_config()['sender'], _email_config()['password'])
        server.sendmail(
            _email_config()['sender'],
            _email_config()['recipients'],
            msg.as_string()
        )

    return True


@app.post("/enviar-correo-pagos-cxc")
async def enviar_correo_pagos_cxc(request: EnviarCorreoCxCRequest, current_user: Usuario = Depends(get_current_user)):
    """Endpoint para enviar correo con los pagos CxC (Anticipos y Cobros a Factura)"""
    try:
        logger.info(f"Enviando correo CxC con {len(request.pagos)} pagos")

        pagos_dict = [p.dict() for p in request.pagos]
        resumen_dict = request.resumen.dict()

        enviar_correo_pagos_cxc_adjunto(pagos_dict, resumen_dict)

        logger.info(f"Correo CxC enviado exitosamente a {_email_config()['recipients']}")

        return {
            "success": True,
            "message": f"Correo CxC enviado: {resumen_dict['total_anticipos']} anticipos, {resumen_dict['total_cobros_factura']} cobros a factura"
        }
    except smtplib.SMTPAuthenticationError:
        logger.error("Error de autenticación SMTP")
        raise HTTPException(status_code=500, detail="Error de autenticación con el servidor de correo")
    except smtplib.SMTPException as e:
        logger.error(f"Error SMTP: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error SMTP: {str(e)}")
    except Exception as e:
        logger.error(f"Error enviando correo CxC: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error enviando correo: {str(e)}")


# ========== ENDPOINTS PORTAL DE SEGUIMIENTO ==========

@app.get("/seguimiento/casos")
def listar_casos(
    estado: str = None,
    tipo: str = None,
    en_proceso: bool = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Lista todos los casos de seguimiento con filtros opcionales."""
    try:
        query = db.query(CasoSeguimiento)

        if estado and estado != 'todos':
            query = query.filter(CasoSeguimiento.estado_seguimiento == estado)

        if tipo and tipo != 'todos':
            query = query.filter(CasoSeguimiento.tipo_problema == tipo)

        if en_proceso is not None:
            query = query.filter(CasoSeguimiento.en_ultimo_proceso == en_proceso)

        casos = query.order_by(CasoSeguimiento.fecha_ultima_actualizacion.desc()).all()
        logger.info(f"Obtenidos {len(casos)} casos")

        return [caso.to_dict() for caso in casos]

    except Exception as e:
        logger.error(f"Error listando casos: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/seguimiento/casos/{referencia}")
def obtener_caso(referencia: str, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """Obtiene un caso específico por referencia."""
    caso = db.query(CasoSeguimiento).filter(CasoSeguimiento.referencia == referencia).first()

    if not caso:
        raise HTTPException(status_code=404, detail="Caso no encontrado")

    return caso.to_dict()

@app.post("/seguimiento/casos")
def crear_caso(caso_data: CasoSeguimientoCreate, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """Crea un nuevo caso de seguimiento."""
    try:
        # Verificar si ya existe
        existente = db.query(CasoSeguimiento).filter(
            CasoSeguimiento.referencia == caso_data.referencia
        ).first()

        if existente:
            raise HTTPException(status_code=400, detail="El caso ya existe")

        caso = CasoSeguimiento(
            referencia=caso_data.referencia,
            tipo_problema=caso_data.tipoProblema,
            estado_seguimiento=caso_data.estadoSeguimiento,
            monto_neto=caso_data.montoNeto,
            monto_bruto=caso_data.montoBruto,
            costos=caso_data.costos,
            notas=caso_data.notas,
            en_ultimo_proceso=caso_data.enUltimoProceso,
            datos_orden=caso_data.datosOrden,
            historial=caso_data.historial
        )

        db.add(caso)
        db.commit()
        db.refresh(caso)

        logger.info(f"Caso creado: {caso_data.referencia}")
        return caso.to_dict()

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error creando caso: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.put("/seguimiento/casos/{referencia}")
def actualizar_caso(referencia: str, caso_data: CasoSeguimientoUpdate, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """Actualiza un caso existente."""
    try:
        caso = db.query(CasoSeguimiento).filter(CasoSeguimiento.referencia == referencia).first()

        if not caso:
            raise HTTPException(status_code=404, detail="Caso no encontrado")

        # Actualizar solo los campos proporcionados
        if caso_data.tipoProblema is not None:
            caso.tipo_problema = caso_data.tipoProblema
        if caso_data.estadoSeguimiento is not None:
            # No revertir un caso ya resuelto salvo que se envíe explícitamente 'resuelto'
            if caso.estado_seguimiento != 'resuelto' or caso_data.estadoSeguimiento == 'resuelto':
                caso.estado_seguimiento = caso_data.estadoSeguimiento
        if caso_data.montoNeto is not None:
            caso.monto_neto = caso_data.montoNeto
        if caso_data.montoBruto is not None:
            caso.monto_bruto = caso_data.montoBruto
        if caso_data.costos is not None:
            caso.costos = caso_data.costos
        if caso_data.notas is not None:
            caso.notas = caso_data.notas
        if caso_data.enUltimoProceso is not None:
            caso.en_ultimo_proceso = caso_data.enUltimoProceso
        if caso_data.datosOrden is not None:
            caso.datos_orden = caso_data.datosOrden
        if caso_data.historial is not None:
            caso.historial = caso_data.historial

        caso.fecha_ultima_actualizacion = datetime.utcnow()

        db.commit()
        db.refresh(caso)

        logger.info(f"Caso actualizado: {referencia}")
        return caso.to_dict()

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error actualizando caso: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.post("/seguimiento/sincronizar")
def sincronizar_casos(request: SincronizarCasosRequest, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """
    Sincroniza casos desde el frontend.
    - Crea casos nuevos
    - Actualiza casos existentes (conserva notas y estado)
    - Marca casos no presentes como fuera de proceso
    """
    try:
        referencias_actuales = set()
        casos_nuevos = 0
        casos_actualizados = 0
        ahora = datetime.utcnow()

        for caso_data in request.casos:
            referencias_actuales.add(caso_data.referencia)

            caso_existente = db.query(CasoSeguimiento).filter(
                CasoSeguimiento.referencia == caso_data.referencia
            ).first()

            if caso_existente:
                # Si el caso ya está resuelto en BD, no modificar estado ni datos operativos
                if caso_existente.estado_seguimiento == 'resuelto':
                    caso_existente.en_ultimo_proceso = caso_data.enUltimoProceso
                    caso_existente.fecha_ultima_actualizacion = ahora
                    casos_actualizados += 1
                    continue

                # Actualizar datos pero conservar notas y estado de seguimiento
                cambios = []

                if caso_existente.monto_neto != caso_data.montoNeto:
                    cambios.append(f"Monto: ${caso_existente.monto_neto:.2f} → ${caso_data.montoNeto:.2f}")
                    caso_existente.monto_neto = caso_data.montoNeto

                if caso_existente.tipo_problema != caso_data.tipoProblema:
                    cambios.append(f"Tipo: {caso_existente.tipo_problema} → {caso_data.tipoProblema}")
                    caso_existente.tipo_problema = caso_data.tipoProblema

                # Actualizar estado solo si el frontend envía uno distinto al actual
                if caso_data.estadoSeguimiento and caso_data.estadoSeguimiento != caso_existente.estado_seguimiento:
                    caso_existente.estado_seguimiento = caso_data.estadoSeguimiento

                caso_existente.monto_bruto = caso_data.montoBruto
                caso_existente.costos = caso_data.costos
                caso_existente.en_ultimo_proceso = caso_data.enUltimoProceso
                caso_existente.datos_orden = caso_data.datosOrden
                caso_existente.fecha_ultima_actualizacion = ahora

                if cambios:
                    historial = caso_existente.historial or []
                    historial.append({
                        "fecha": ahora.isoformat(),
                        "tipo": "actualizacion_datos",
                        "detalle": "; ".join(cambios)
                    })
                    caso_existente.historial = historial

                casos_actualizados += 1
            else:
                # Crear nuevo caso
                historial_inicial = [{
                    "fecha": ahora.isoformat(),
                    "tipo": "creacion",
                    "detalle": f"Caso creado - {caso_data.tipoProblema}"
                }]

                caso_nuevo = CasoSeguimiento(
                    referencia=caso_data.referencia,
                    tipo_problema=caso_data.tipoProblema,
                    estado_seguimiento="pendiente",
                    monto_neto=caso_data.montoNeto,
                    monto_bruto=caso_data.montoBruto,
                    costos=caso_data.costos,
                    notas="",
                    en_ultimo_proceso=True,
                    datos_orden=caso_data.datosOrden,
                    historial=historial_inicial,
                    fecha_deteccion=ahora,
                    fecha_ultima_actualizacion=ahora
                )
                db.add(caso_nuevo)
                casos_nuevos += 1

        # Marcar casos no presentes como fuera de proceso
        # EXCLUYE casos Walmart (referencia WM-*) — se gestionan por su propio proceso
        casos_no_presentes = db.query(CasoSeguimiento).filter(
            CasoSeguimiento.en_ultimo_proceso == True,
            ~CasoSeguimiento.referencia.in_(referencias_actuales),
            ~CasoSeguimiento.referencia.like("WM-%")
        ).all()

        for caso in casos_no_presentes:
            caso.en_ultimo_proceso = False
            historial = caso.historial or []
            historial.append({
                "fecha": ahora.isoformat(),
                "tipo": "no_en_proceso",
                "detalle": "Orden no aparece en el último procesamiento"
            })
            caso.historial = historial

        db.commit()

        logger.info(f"Sincronización completada: {casos_nuevos} nuevos, {casos_actualizados} actualizados, {len(casos_no_presentes)} marcados fuera de proceso")

        return {
            "success": True,
            "casosNuevos": casos_nuevos,
            "casosActualizados": casos_actualizados,
            "casosFueraProceso": len(casos_no_presentes),
            "totalCasos": len(referencias_actuales)
        }

    except Exception as e:
        db.rollback()
        logger.error(f"Error sincronizando casos: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/seguimiento/estadisticas")
def obtener_estadisticas(db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """Obtiene estadísticas del portal de seguimiento."""
    try:
        total = db.query(CasoSeguimiento).count()
        pendientes = db.query(CasoSeguimiento).filter(CasoSeguimiento.estado_seguimiento == 'pendiente').count()
        revision = db.query(CasoSeguimiento).filter(CasoSeguimiento.estado_seguimiento == 'revision').count()
        resueltos = db.query(CasoSeguimiento).filter(CasoSeguimiento.estado_seguimiento == 'resuelto').count()

        return {
            "total": total,
            "pendientes": pendientes,
            "revision": revision,
            "resueltos": resueltos
        }

    except Exception as e:
        logger.error(f"Error obteniendo estadísticas: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")



@app.delete("/seguimiento/casos/{referencia}")
def eliminar_caso(referencia: str, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """Elimina un caso de seguimiento."""
    try:
        caso = db.query(CasoSeguimiento).filter(CasoSeguimiento.referencia == referencia).first()

        if not caso:
            raise HTTPException(status_code=404, detail="Caso no encontrado")

        db.delete(caso)
        db.commit()

        logger.info(f"Caso eliminado: {referencia}")
        return {"success": True, "message": f"Caso {referencia} eliminado"}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error eliminando caso: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ========== ENDPOINTS HISTÓRICO DE CONCILIACIONES ==========

@app.post("/historico/conciliaciones")
def guardar_conciliacion(request: GuardarConciliacionRequest, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """Guarda una conciliación completa (header + órdenes)."""
    try:
        conciliacion = HistoricoConciliacion(
            fecha=datetime.utcnow(),
            total_ordenes=request.totalOrdenes,
            coincidencias=request.coincidencias,
            casos_creados=request.casosCreados,
            ordenes_validadas=request.ordenesValidadas,
            monto_estado_cuenta=request.montoEstadoCuenta,
            total_ingresos_bruto=request.totalIngresosBruto,
            total_costos=request.totalCostos,
            total_neto_real=request.totalNetoReal,
            resumen_por_estatus=request.resumenPorEstatus,
            notas=request.notas
        )
        db.add(conciliacion)
        db.flush()

        if request.ordenes:
            ordenes_db = [
                HistoricoOrden(
                    conciliacion_id=conciliacion.id,
                    referencia_erp=o.referenciaERP,
                    estatus=o.estatus,
                    datos_orden=o.datosOrden
                )
                for o in request.ordenes
            ]
            db.bulk_save_objects(ordenes_db)

        db.commit()
        db.refresh(conciliacion)

        logger.info(f"Conciliación guardada: id={conciliacion.id}, {request.totalOrdenes} órdenes")
        return conciliacion.to_dict()

    except Exception as e:
        db.rollback()
        logger.error(f"Error guardando conciliación: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/historico/conciliaciones")
def listar_conciliaciones(limit: int = 50, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """Lista headers de conciliaciones (sin órdenes), orden desc por fecha."""
    try:
        conciliaciones = db.query(HistoricoConciliacion)\
            .order_by(HistoricoConciliacion.fecha.desc())\
            .limit(limit)\
            .all()
        return [c.to_dict() for c in conciliaciones]
    except Exception as e:
        logger.error(f"Error listando conciliaciones: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/historico/conciliaciones/{conciliacion_id}")
def obtener_conciliacion(conciliacion_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """Obtiene header + todas las órdenes de una conciliación."""
    conciliacion = db.query(HistoricoConciliacion).filter(
        HistoricoConciliacion.id == conciliacion_id
    ).first()

    if not conciliacion:
        raise HTTPException(status_code=404, detail="Conciliación no encontrada")

    ordenes = db.query(HistoricoOrden).filter(
        HistoricoOrden.conciliacion_id == conciliacion_id
    ).all()

    result = conciliacion.to_dict()
    result["ordenes"] = [o.to_dict() for o in ordenes]
    return result


@app.delete("/historico/conciliaciones/{conciliacion_id}")
def eliminar_conciliacion(conciliacion_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """Elimina una conciliación y sus órdenes."""
    try:
        conciliacion = db.query(HistoricoConciliacion).filter(
            HistoricoConciliacion.id == conciliacion_id
        ).first()

        if not conciliacion:
            raise HTTPException(status_code=404, detail="Conciliación no encontrada")

        db.query(HistoricoOrden).filter(
            HistoricoOrden.conciliacion_id == conciliacion_id
        ).delete()
        db.delete(conciliacion)
        db.commit()

        logger.info(f"Conciliación eliminada: id={conciliacion_id}")
        return {"success": True, "message": f"Conciliación {conciliacion_id} eliminada"}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error eliminando conciliación: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/historico/comparar")
def comparar_conciliaciones(ids: str, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """Retorna headers de las conciliaciones indicadas para comparación."""
    try:
        id_list = [int(x.strip()) for x in ids.split(",") if x.strip()]
    except ValueError:
        raise HTTPException(status_code=400, detail="IDs inválidos")

    if len(id_list) < 2:
        raise HTTPException(status_code=400, detail="Se requieren al menos 2 IDs")

    conciliaciones = db.query(HistoricoConciliacion).filter(
        HistoricoConciliacion.id.in_(id_list)
    ).order_by(HistoricoConciliacion.fecha.asc()).all()

    if len(conciliaciones) < 2:
        raise HTTPException(status_code=404, detail="No se encontraron suficientes conciliaciones")

    return [c.to_dict() for c in conciliaciones]


# ========== WALMART MX – CONCILIACIÓN ==========

def _sync_casos_walmart(casos_data: list, db: Session) -> dict:
    """Crea o actualiza casos de devolución Walmart en la BD. Devuelve conteos."""
    ahora  = datetime.utcnow()
    nuevos = actualizados = 0
    for cd in casos_data:
        existente = db.query(CasoSeguimiento).filter(
            CasoSeguimiento.referencia == cd["referencia"]
        ).first()
        if existente:
            existente.monto_neto  = cd["montoNeto"]
            existente.monto_bruto = cd["montoBruto"]
            existente.datos_orden = cd["datosOrden"]
            existente.en_ultimo_proceso = True
            existente.fecha_ultima_actualizacion = ahora
            hist = list(existente.historial or [])
            hist.append({"fecha": ahora.isoformat(), "tipo": "sistema",
                         "detalle": "Actualizado en conciliación Walmart"})
            existente.historial = hist
            actualizados += 1
        else:
            db.add(CasoSeguimiento(
                referencia=cd["referencia"],
                tipo_problema=cd["tipoProblema"],
                estado_seguimiento="pendiente",
                monto_neto=cd["montoNeto"],
                monto_bruto=cd["montoBruto"],
                costos=0.0,
                notas="",
                en_ultimo_proceso=True,
                datos_orden=cd["datosOrden"],
                historial=[{"fecha": ahora.isoformat(), "tipo": "creacion",
                            "detalle": f"Caso creado – {cd['tipoProblema']}"}],
                fecha_deteccion=ahora,
                fecha_ultima_actualizacion=ahora,
            ))
            nuevos += 1
    db.commit()
    return {"casos_nuevos": nuevos, "casos_actualizados": actualizados}


def _extraer_notas_credito(xlsx_path: Path, cuenta_default: str = "") -> list:
    """Lee la hoja de Notas de Crédito del Excel generado y retorna lista de casos."""
    wb = load_workbook(xlsx_path, read_only=True)

    # Buscar la hoja: consolidado → "Notas de Crédito", individual → "🔄 Notas de Crédito"
    nombre_hoja = next(
        (s for s in wb.sheetnames if "Notas" in s and "r" in s.lower()),
        None
    )
    if not nombre_hoja:
        wb.close()
        return []

    ws      = wb[nombre_hoja]
    headers = None
    casos   = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        if headers is None:
            # Fila 4 = encabezados
            headers = [str(c).strip() if c is not None else "" for c in row]
            continue
        if not any(row):
            continue
        rd = dict(zip(headers, row))

        cuenta    = str(rd.get("Cuenta", cuenta_default) or cuenta_default).strip()
        orderline = str(rd.get("Orderline Number", "") or "").strip()
        amount    = abs(float(rd.get("Amount", 0) or 0))
        fecha     = str(rd.get("Payment Date", "") or "")
        erp_id    = str(rd.get("ERP_ID", "") or "")
        fulfilled = str(rd.get("FulfilledBy", "") or "")

        if not orderline:
            continue

        casos.append({
            "referencia":  f"WM-{cuenta.upper()}-{orderline}",
            "tipoProblema": f"DEVOLUCION_WALMART_{cuenta.upper()}",
            "montoNeto":   amount,
            "montoBruto":  amount,
            "datosOrden":  {
                "cuenta": cuenta,
                "orderlineNumber": orderline,
                "paymentDate": fecha,
                "erpId": erp_id,
                "fulfilledBy": fulfilled,
                "plataforma": "Walmart MX",
            },
        })

    wb.close()
    return casos


@app.post("/walmart/conciliar")
async def walmart_conciliar(
    atlas_csv:  UploadFile = File(...),
    spring_csv: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Conciliación consolidada Walmart MX — Atlas + Spring."""
    run_id      = uuid.uuid4().hex[:8]
    atlas_path  = CONCILIACION_DIR / f"_tmp_atlas_{run_id}.csv"
    spring_path = CONCILIACION_DIR / f"_tmp_spring_{run_id}.csv"
    output_name = f"consolidado_walmart_{run_id}.xlsx"
    output_path = CONCILIACION_DIR / output_name

    try:
        atlas_path.write_bytes(await atlas_csv.read())
        spring_path.write_bytes(await spring_csv.read())

        result = subprocess.run(
            [sys.executable, "run_consolidado.py",
             "--atlas",  str(atlas_path),
             "--spring", str(spring_path),
             "--output", str(output_path)],
            cwd=str(CONCILIACION_DIR),
            capture_output=True, text=True, timeout=300,
        )
        if result.returncode != 0:
            raise HTTPException(status_code=500,
                detail=f"Error en conciliación:\n{result.stderr[-800:]}")

        # Extraer Notas de Crédito y sincronizar casos
        casos  = _extraer_notas_credito(output_path)
        counts = _sync_casos_walmart(casos, db)

        # Contar registros totales por cuenta (hoja A Pagar Proveedor, col 0 = Cuenta)
        atlas_regs = spring_regs = 0
        wb2 = load_workbook(output_path, read_only=True)
        if "A Pagar Proveedor" in wb2.sheetnames:
            for row in wb2["A Pagar Proveedor"].iter_rows(min_row=5, values_only=True):
                if not any(row): continue
                cv = str(row[0] or "").strip()
                if cv == "Atlas":  atlas_regs  += 1
                elif cv == "Spring": spring_regs += 1
        wb2.close()

        logger.info(f"Walmart consolidado run={run_id}: atlas={atlas_regs} spring={spring_regs} {counts}")
        return {
            "success": True,
            "archivo": output_name,
            "resumen": {
                "atlas":  atlas_regs,
                "spring": spring_regs,
                **counts,
            },
        }

    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=504, detail="El proceso excedió 5 minutos")
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error walmart_conciliar: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        for p in (atlas_path, spring_path):
            p.unlink(missing_ok=True)


@app.post("/walmart/conciliar-individual")
async def walmart_conciliar_individual(
    csv_file: UploadFile = File(...),
    cuenta: str = "atlas",
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Conciliación individual Walmart MX — una cuenta."""
    if cuenta not in ("atlas", "spring"):
        raise HTTPException(status_code=400, detail="cuenta debe ser 'atlas' o 'spring'")

    run_id      = uuid.uuid4().hex[:8]
    csv_path    = CONCILIACION_DIR / f"_tmp_{cuenta}_{run_id}.csv"
    output_name = f"conciliacion_{cuenta}_{run_id}.xlsx"
    output_path = CONCILIACION_DIR / output_name

    try:
        csv_path.write_bytes(await csv_file.read())

        result = subprocess.run(
            [sys.executable, "run_conciliacion.py",
             str(csv_path), "--cuenta", cuenta, "--output", str(output_path)],
            cwd=str(CONCILIACION_DIR),
            capture_output=True, text=True, timeout=300,
        )
        if result.returncode != 0:
            raise HTTPException(status_code=500,
                detail=f"Error en conciliación:\n{result.stderr[-800:]}")

        casos  = _extraer_notas_credito(output_path, cuenta_default=cuenta.capitalize())
        counts = _sync_casos_walmart(casos, db)

        logger.info(f"Walmart individual run={run_id} cuenta={cuenta}: {counts}")
        return {
            "success": True,
            "archivo": output_name,
            "resumen": {cuenta: len(casos), **counts},
        }

    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=504, detail="El proceso excedió 5 minutos")
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error walmart_conciliar_individual: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        csv_path.unlink(missing_ok=True)


@app.get("/walmart/descargar/{filename}")
async def walmart_descargar(
    filename: str,
    current_user: Usuario = Depends(get_current_user),
):
    """Descarga el Excel de conciliación Walmart generado."""
    if not re.match(r'^(consolidado_walmart|conciliacion_(atlas|spring))_[a-f0-9]{8}\.xlsx$', filename):
        raise HTTPException(status_code=400, detail="Nombre de archivo inválido")
    file_path = CONCILIACION_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado o expirado")
    return FileResponse(
        path=str(file_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=settings.api_port,
        reload=True
    )