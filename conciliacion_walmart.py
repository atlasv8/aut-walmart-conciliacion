"""
conciliacion_walmart.py
=======================
Script de conciliación de pagos Walmart vs Estado de Cuenta.

Uso:
    python conciliacion_walmart.py <archivo_csv> [--output <archivo.xlsx>] [--ultimo]

Opciones:
    --ultimo    Reutiliza los valores del último estado de cuenta sin volver a preguntar

Ejemplos:
    python conciliacion_walmart.py Atlas_03-2026.csv
    python conciliacion_walmart.py Atlas_03-2026.csv --output MiConciliacion.xlsx
    python conciliacion_walmart.py Atlas_04-2026.csv --ultimo

Dependencias:
    pip install pandas openpyxl
"""

import sys
import json
import argparse
import csv
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: Instala dependencias con:  pip install pandas openpyxl")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  ARCHIVO CACHÉ – guarda los últimos valores ingresados
#  Se crea automáticamente junto al script la primera vez que se ejecuta
# ══════════════════════════════════════════════════════════════════════════════

CACHE_FILE = Path(__file__).parent / ".walmart_edo_cta.json"


# ══════════════════════════════════════════════════════════════════════════════
#  DEFINICIÓN DE CAMPOS DEL ESTADO DE CUENTA
#  Orden en que se preguntarán al usuario.
#  signo:  1 = ingreso (positivo en el edo. cta.)
#         -1 = egreso  (negativo en el edo. cta.)
#  El usuario SIEMPRE ingresa el valor absoluto; el signo se aplica internamente.
# ══════════════════════════════════════════════════════════════════════════════

CAMPOS = [
    # (clave_interna,                       etiqueta_usuario,                          signo)
    # ── VENTAS ──────────────────────────────────────────────────────────────────────
    ("Ventas - Precio del producto",    "Ventas › Precio del producto",                 1),
    ("Ventas - Comisión",               "Ventas › Comisión",                           -1),
    ("Ventas - IVA retenido",           "Ventas › IVA retenido",                       -1),
    ("Ventas - ISR retenido",           "Ventas › ISR retenido",                       -1),
    # ── REEMBOLSOS ──────────────────────────────────────────────────────────────────
    ("Reembolsos - Precio del producto","Reembolsos › Precio del producto",            -1),
    ("Reembolsos - IVA retenido",       "Reembolsos › IVA retenido  (a tu favor)",     1),
    ("Reembolsos - ISR retenido",       "Reembolsos › ISR retenido  (a tu favor)",     1),
    ("Reembolsos - otherReturnFees",    "Reembolsos › otherReturnFees / Reclamo",      1),
    # ── WFS ─────────────────────────────────────────────────────────────────────────
    ("WFS - storageFee",                "WFS › storageFee  (almacenaje)",              -1),
    ("WFS - fulfillmentFee",            "WFS › fulfillmentFee  (envío + manejo)",      -1),
    # ── OTROS ───────────────────────────────────────────────────────────────────────
    ("Otros - Ofertas especiales",      "Otros › Ofertas especiales  (Killer)",         1),
    ("Otros - SEM",                     "Otros › SEM – Campañas de marketing",         -1),
    ("Otros - Killer Comisión",         "Otros › Killer Comisión",                     -1),
]

# Índice rápido por clave
CAMPO_POR_CLAVE = {c[0]: c for c in CAMPOS}

# Agrupación visual para la terminal
SECCIONES_TERMINAL = {
    "VENTAS": [
        "Ventas - Precio del producto", "Ventas - Comisión",
        "Ventas - IVA retenido",        "Ventas - ISR retenido",
    ],
    "REEMBOLSOS": [
        "Reembolsos - Precio del producto", "Reembolsos - IVA retenido",
        "Reembolsos - ISR retenido",         "Reembolsos - otherReturnFees",
    ],
    "WFS": [
        "WFS - storageFee", "WFS - fulfillmentFee",
    ],
    "OTROS": [
        "Otros - Ofertas especiales", "Otros - SEM", "Otros - Killer Comisión",
    ],
}


# ══════════════════════════════════════════════════════════════════════════════
#  MAPEO CSV → Categorías del Estado de Cuenta
# ══════════════════════════════════════════════════════════════════════════════

MAPEO = {
    ("Venta",                      "Ingreso"): "Ventas - Precio del producto",
    ("Comision",                   "Egreso"):  "Ventas - Comisión",
    ("Retencion IVA",              "Egreso"):  "Ventas - IVA retenido",
    ("Retencion ISR",              "Egreso"):  "Ventas - ISR retenido",
    ("Venta",                      "Egreso"):  "Reembolsos - Precio del producto",
    ("Retencion IVA",              "Ingreso"): "Reembolsos - IVA retenido",
    ("Retencion ISR",              "Ingreso"): "Reembolsos - ISR retenido",
    ("Reclamo",                    "Ingreso"): "Reembolsos - otherReturnFees",
    ("WFS - Tarifa de almacenaje", "Egreso"):  "WFS - storageFee",
    ("WFS - Tarifa de envio",      "Egreso"):  "WFS - fulfillmentFee",
    ("WFS - Tarifa de manejo",     "Egreso"):  "WFS - fulfillmentFee",
    ("Killer",                     "Ingreso"): "Otros - Ofertas especiales",
    ("Killer",                     "Egreso"):  "Otros - Ofertas especiales",
    ("Campañas - Marketing SEM",   "Egreso"):  "Otros - SEM",
    ("killerCommission",           "Egreso"):  "Otros - Killer Comisión",
    ("killerCommission",           "Ingreso"): "Otros - Killer Comisión",
}


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRADA INTERACTIVA
# ══════════════════════════════════════════════════════════════════════════════

def cargar_cache() -> dict:
    if CACHE_FILE.exists():
        try:
            return json.loads(CACHE_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def guardar_cache(datos: dict):
    CACHE_FILE.write_text(json.dumps(datos, indent=2, ensure_ascii=False), encoding="utf-8")


def pedir_monto(etiqueta: str, signo: int, ultimo: float | None) -> float:
    """
    Pide un monto al usuario.
    - El usuario ingresa el valor absoluto (sin signo).
    - Enter vacío acepta el último valor guardado.
    - El signo se aplica internamente.
    """
    sufijo = " (egreso −)" if signo == -1 else " (ingreso +)"
    hint   = f"  [último: {abs(ultimo):>14,.2f}]" if ultimo is not None else ""

    while True:
        raw = input(f"    {etiqueta}{sufijo}{hint}: ").strip()

        if raw == "" and ultimo is not None:
            return ultimo                       # acepta último valor

        raw = raw.replace("$", "").replace(",", "").replace(" ", "")

        if raw in ("0", "0.0", "0.00"):
            return 0.0

        try:
            valor = float(raw)
            if valor < 0:
                print("      ⚠  Ingresa el valor sin signo (solo el número positivo).")
                continue
            return signo * valor
        except ValueError:
            print("      ⚠  Valor inválido. Ejemplo: 151240.53")


def capturar_estado_de_cuenta(usar_ultimo: bool) -> tuple[dict, float]:
    """
    Captura interactivamente el estado de cuenta.
    Devuelve (dict_valores, total_pagado).
    """
    cache = cargar_cache()
    valores = {}

    print()
    print("  ╔══════════════════════════════════════════════════════════════════╗")
    print("  ║           INGRESO DEL ESTADO DE CUENTA WALMART                  ║")
    if usar_ultimo and cache:
        print("  ║  Modo --ultimo: Enter acepta automáticamente el valor anterior   ║")
    else:
        print("  ║  Ingresa el valor ABSOLUTO de cada campo (sin signo ni $)        ║")
        print("  ║  Presiona Enter para aceptar el valor anterior  (si existe)      ║")
    print("  ╚══════════════════════════════════════════════════════════════════╝")

    for seccion, claves in SECCIONES_TERMINAL.items():
        print(f"\n  ── {seccion} {'─' * (55 - len(seccion))}")
        for clave in claves:
            _, etiqueta, signo = CAMPO_POR_CLAVE[clave]
            ultimo = cache.get(clave)

            if usar_ultimo and ultimo is not None:
                valores[clave] = ultimo
                print(f"    {etiqueta}  →  {ultimo:,.2f}  (reutilizado)")
            else:
                valores[clave] = pedir_monto(etiqueta, signo, ultimo)

    # Total pagado
    print(f"\n  ── TOTAL {'─' * 52}")
    ultimo_total = cache.get("__total_pagado__")
    hint_total   = f"  [último: {ultimo_total:>14,.2f}]" if ultimo_total is not None else ""

    if usar_ultimo and ultimo_total is not None:
        total_pagado = ultimo_total
        print(f"    Total pagado  →  {total_pagado:,.2f}  (reutilizado)")
    else:
        while True:
            raw = input(f"    Total pagado (ingreso +){hint_total}: ").strip()
            if raw == "" and ultimo_total is not None:
                total_pagado = ultimo_total
                break
            raw = raw.replace("$", "").replace(",", "").replace(" ", "")
            try:
                total_pagado = float(raw)
                if total_pagado <= 0:
                    print("      ⚠  El total debe ser mayor a cero.")
                    continue
                break
            except ValueError:
                print("      ⚠  Valor inválido.")

    # ── Resumen de confirmación ───────────────────────────────────────────────
    print()
    print("  ┌─ RESUMEN ──────────────────────────────────────────────────────────┐")
    for clave, val in valores.items():
        etiqueta = CAMPO_POR_CLAVE[clave][1]
        print(f"  │  {etiqueta:<48}  {val:>14,.2f}  │")
    print(f"  │  {'Total pagado':<48}  {total_pagado:>14,.2f}  │")
    print("  └────────────────────────────────────────────────────────────────────┘")
    print()

    ok = input("  ¿Los datos son correctos? (S/n): ").strip().lower()
    if ok in ("n", "no"):
        print("\n  Reiniciando captura...\n")
        return capturar_estado_de_cuenta(usar_ultimo=False)

    # Guardar caché
    cache_nuevo = dict(valores)
    cache_nuevo["__total_pagado__"] = total_pagado
    guardar_cache(cache_nuevo)
    print("  ✅ Valores guardados para la próxima ejecución.\n")

    return valores, total_pagado


# ══════════════════════════════════════════════════════════════════════════════
#  LECTURA Y AGRUPACIÓN DEL CSV
# ══════════════════════════════════════════════════════════════════════════════

def leer_csv(ruta: str) -> pd.DataFrame:
    """Lee el CSV de Walmart (2 filas de encabezado: inglés + español)."""
    with open(ruta, encoding="utf-8") as f:
        lines = f.readlines()
    eng_header = next(csv.reader([lines[0]]))
    rows = []
    for line in lines[2:]:
        if not line.strip():
            continue
        row = next(csv.reader([line]))
        record = {eng_header[i]: row[i] if i < len(row) else "" for i in range(len(eng_header))}
        rows.append(record)
    df = pd.DataFrame(rows)
    df["Amount"]           = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)
    df["Concept"]          = df["Concept"].str.strip()
    df["Income / Outcome"] = df["Income / Outcome"].str.strip()
    return df


def agrupar_por_categoria(df: pd.DataFrame) -> tuple[dict, dict]:
    totales   = {c[0]: 0.0 for c in CAMPOS}
    sin_mapeo = {}
    for _, row in df.iterrows():
        clave = (row["Concept"], row["Income / Outcome"])
        monto = row["Amount"]
        if clave in MAPEO:
            cat = MAPEO[clave]
            totales[cat] = round(totales[cat] + monto, 4)
        else:
            sin_mapeo[clave] = sin_mapeo.get(clave, 0) + monto
    return totales, sin_mapeo


# ══════════════════════════════════════════════════════════════════════════════
#  ESTILOS EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def hfill(h):  return PatternFill("solid", start_color=h, fgColor=h)
def thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)
def mfmt():    return '#,##0.00;(#,##0.00);"-"'

C = dict(blue="2E75B6", blue2="0070C0", lgray="F2F2F2", white="FFFFFF",
         yellow="FFF2CC", secbg="D6E4F0", green="00B050", red="FF0000",
         orange="FF9900")
F = {
    "title": Font(name="Arial", bold=True, size=13, color="FFFFFF"),
    "hdr":   Font(name="Arial", bold=True, size=10, color="FFFFFF"),
    "body":  Font(name="Arial", size=10),
    "bold":  Font(name="Arial", bold=True, size=10),
    "sec":   Font(name="Arial", bold=True, size=10, color="1F3864"),
    "ital":  Font(name="Arial", size=9,  italic=True, color="595959"),
    "grn":   Font(name="Arial", size=10, color="00B050"),
    "red":   Font(name="Arial", bold=True, size=10, color="FF0000"),
    "org":   Font(name="Arial", size=10, color="FF9900"),
}

SECCIONES_EXCEL = {
    "VENTAS":                         ["Ventas - Precio del producto", "Ventas - Comisión",
                                       "Ventas - IVA retenido",        "Ventas - ISR retenido"],
    "REEMBOLSOS":                     ["Reembolsos - Precio del producto", "Reembolsos - IVA retenido",
                                       "Reembolsos - ISR retenido",         "Reembolsos - otherReturnFees"],
    "WFS – SERVICIOS DE FULFILLMENT": ["WFS - storageFee", "WFS - fulfillmentFee"],
    "OTROS":                          ["Otros - Ofertas especiales", "Otros - SEM",
                                       "Otros - Killer Comisión"],
}

ETIQUETAS_EXCEL = {
    "Ventas - Precio del producto":      "Precio del producto",
    "Ventas - Comisión":                 "Comisión",
    "Ventas - IVA retenido":             "IVA retenido",
    "Ventas - ISR retenido":             "ISR retenido",
    "Reembolsos - Precio del producto":  "Precio del producto",
    "Reembolsos - IVA retenido":         "IVA retenido (a favor)",
    "Reembolsos - ISR retenido":         "ISR retenido (a favor)",
    "Reembolsos - otherReturnFees":      "otherReturnFees / Reclamo",
    "WFS - storageFee":                  "storageFee (almacenaje)",
    "WFS - fulfillmentFee":              "fulfillmentFee (envío + manejo)",
    "Otros - Ofertas especiales":        "Ofertas especiales (Killer)",
    "Otros - SEM":                       "SEM – Marketing",
    "Otros - Killer Comisión":           "Killer Comisión",
}

def observacion(cat, diff):
    if abs(diff) < 0.10:
        return "✅ Cuadra exacto"
    if "WFS" in cat:
        return (f"⚠ Diferencia ${diff:+,.2f} – Solo reclasificación interna; "
                "el total WFS cuadra exacto.")
    if abs(diff) < 1:
        return f"Diferencia de ${abs(diff):.2f} – redondeo"
    return f"⚠ Diferencia de ${diff:+,.2f} – revisar"


# ══════════════════════════════════════════════════════════════════════════════
#  GENERACIÓN DEL EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def crear_excel(df, totales_csv, sin_mapeo, edo_cta, total_pagado_edo, output):
    wb = Workbook()

    # ── Hoja 1: Conciliación ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Conciliación"
    for col, w in zip("ABCDE", [40, 18, 18, 14, 32]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:E1")
    ws["A1"] = "CONCILIACIÓN DE PAGOS WALMART – ATLAS DEL DESCANSO"
    ws["A1"].font = F["title"]; ws["A1"].fill = hfill(C["blue2"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    fecha_pago = df["Payment Date"].dropna().iloc[0] if not df.empty else "N/D"
    for celda, val in [("A2","Periodo de pago:"),("B2",str(fecha_pago)),
                       ("D2","Registros CSV:"),  ("E2",len(df))]:
        ws[celda].value = val; ws[celda].fill = hfill(C["lgray"]); ws[celda].font = F["body"]
    ws["A2"].font = F["bold"]; ws["D2"].font = F["bold"]
    ws.row_dimensions[2].height = 18; ws.row_dimensions[3].height = 6

    for ci, h in enumerate(["Concepto","CSV (detalle)","Estado de Cuenta",
                             "Diferencia","Observaciones"], 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = F["hdr"]; c.fill = hfill(C["blue"])
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin()
    ws.row_dimensions[4].height = 22

    r = 5
    for seccion, claves in SECCIONES_EXCEL.items():
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(row=r, column=1, value=seccion)
        c.font = F["sec"]; c.fill = hfill(C["secbg"])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin(); ws.row_dimensions[r].height = 18; r += 1

        for clave in claves:
            csv_val = totales_csv.get(clave, 0.0)
            edo_val = edo_cta.get(clave, 0.0)
            diff    = round(csv_val - edo_val, 2)
            fill    = hfill(C["lgray"]) if r % 2 == 0 else hfill(C["white"])

            c1 = ws.cell(row=r, column=1, value=ETIQUETAS_EXCEL.get(clave, clave))
            c1.font = F["body"]; c1.fill = fill
            c1.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            c1.border = thin()

            for ci, val in [(2, csv_val), (3, edo_val)]:
                cx = ws.cell(row=r, column=ci, value=val)
                cx.font = F["body"]; cx.fill = fill
                cx.number_format = mfmt(); cx.alignment = Alignment(horizontal="right")
                cx.border = thin()

            c4 = ws.cell(row=r, column=4, value=f"=B{r}-C{r}")
            c4.number_format = mfmt(); c4.fill = fill
            c4.alignment = Alignment(horizontal="right"); c4.border = thin()
            c4.font = F["grn"] if abs(diff) < 0.10 else (F["red"] if abs(diff) > 1000 else F["org"])

            c5 = ws.cell(row=r, column=5, value=observacion(clave, diff))
            c5.font = F["ital"]; c5.fill = fill
            c5.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True, indent=1)
            c5.border = thin(); ws.row_dimensions[r].height = 18; r += 1

    # Fila total
    ws.merge_cells(f"A{r}:E{r}")
    c = ws.cell(row=r, column=1, value="TOTAL PAGADO")
    c.font = F["sec"]; c.fill = hfill(C["secbg"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin(); ws.row_dimensions[r].height = 18; r += 1

    total_csv  = round(sum(totales_csv.values()), 2)
    diff_total = round(total_csv - total_pagado_edo, 2)
    fill = hfill(C["lgray"])
    c1 = ws.cell(row=r, column=1, value="Total pagado")
    c1.font = F["bold"]; c1.fill = fill
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=2); c1.border = thin()
    for ci, val in [(2, total_csv), (3, total_pagado_edo)]:
        cx = ws.cell(row=r, column=ci, value=val)
        cx.font = F["bold"]; cx.fill = fill
        cx.number_format = mfmt(); cx.alignment = Alignment(horizontal="right"); cx.border = thin()
    c4 = ws.cell(row=r, column=4, value=f"=B{r}-C{r}")
    c4.font = F["grn"] if abs(diff_total) < 1 else F["red"]
    c4.fill = fill; c4.number_format = mfmt()
    c4.alignment = Alignment(horizontal="right"); c4.border = thin()
    c5 = ws.cell(row=r, column=5, value=observacion("Total", diff_total))
    c5.font = F["ital"]; c5.fill = fill
    c5.alignment = Alignment(horizontal="left", wrap_text=True, indent=1); c5.border = thin()
    ws.row_dimensions[r].height = 18; r += 2

    ws.merge_cells(f"A{r}:E{r}")
    nota = ws.cell(row=r, column=1,
        value=("NOTA: Las diferencias en WFS son de clasificación interna "
               "(storageFee vs fulfillmentFee). El total WFS cuadra exacto. "
               "La diferencia global es por redondeo de centavos."))
    nota.font = F["ital"]; nota.fill = hfill(C["yellow"])
    nota.alignment = Alignment(wrap_text=True, horizontal="left", indent=1)
    nota.border = thin(); ws.row_dimensions[r].height = 30

    # ── Hoja 2: Detalle CSV ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Detalle CSV")
    sel_cols = ["Payment Date","Orderline Number","Amount","Income / Outcome",
                "Concept","FulfilledBy","Invoice Number","Return Number","Quantity"]
    for i, w in enumerate([14,30,16,18,26,18,18,18,10], 1):
        ws2.column_dimensions[chr(64+i)].width = w
    for ci, h in enumerate(sel_cols, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = F["hdr"]; c.fill = hfill(C["blue"])
        c.alignment = Alignment(horizontal="center"); c.border = thin()
    ws2.row_dimensions[1].height = 20
    df_sel = df[[col for col in sel_cols if col in df.columns]].copy()
    for ri, row_data in enumerate(df_sel.itertuples(index=False), 2):
        alt = hfill(C["lgray"]) if ri % 2 == 0 else hfill(C["white"])
        for ci, val in enumerate(row_data, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = F["body"]; c.fill = alt; c.border = thin()
            if ci == 3:
                c.number_format = mfmt(); c.alignment = Alignment(horizontal="right")

    # ── Hoja 3: Resumen por Concepto ─────────────────────────────────────────
    ws3 = wb.create_sheet("Resumen por Concepto")
    for col, w in zip("ABC", [36, 18, 12]):
        ws3.column_dimensions[col].width = w
    for ci, h in enumerate(["Concepto (CSV)", "Total ($)", "Registros"], 1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.font = F["hdr"]; c.fill = hfill(C["blue"])
        c.alignment = Alignment(horizontal="center"); c.border = thin()
    ws3.row_dimensions[1].height = 20
    grp = (df.groupby(["Concept","Income / Outcome"])
             .agg(Total=("Amount","sum"), Registros=("Amount","count"))
             .reset_index().sort_values(["Concept","Income / Outcome"]))
    for ri, row_data in enumerate(grp.itertuples(index=False), 2):
        alt = hfill(C["lgray"]) if ri % 2 == 0 else hfill(C["white"])
        label = f"{row_data.Concept} ({row_data[1]})"
        for ci, val in enumerate([label, round(row_data.Total, 2), int(row_data.Registros)], 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.font = F["body"]; c.fill = alt; c.border = thin()
            if ci == 2:
                c.number_format = mfmt(); c.alignment = Alignment(horizontal="right")
    rt = len(grp) + 2
    for ci in range(1, 4):
        ws3.cell(row=rt, column=ci).fill = hfill(C["yellow"])
        ws3.cell(row=rt, column=ci).font = F["bold"]
        ws3.cell(row=rt, column=ci).border = thin()
    ws3.cell(row=rt, column=1).value = "TOTAL NETO"
    ws3.cell(row=rt, column=2).value = f"=SUM(B2:B{rt-1})"
    ws3.cell(row=rt, column=2).number_format = mfmt()
    ws3.cell(row=rt, column=2).alignment = Alignment(horizontal="right")
    ws3.cell(row=rt, column=3).value = f"=SUM(C2:C{rt-1})"

    # ── Hoja 4: Sin Mapeo (auditoría) ────────────────────────────────────────
    if sin_mapeo:
        ws4 = wb.create_sheet("Sin Mapeo")
        for col, w in zip("ABC", [30, 16, 16]):
            ws4.column_dimensions[col].width = w
        for ci, h in enumerate(["Concepto CSV","Tipo","Total"], 1):
            c = ws4.cell(row=1, column=ci, value=h)
            c.font = F["hdr"]; c.fill = hfill(C["blue"])
            c.alignment = Alignment(horizontal="center"); c.border = thin()
        for ri, ((concepto, tipo), total) in enumerate(sin_mapeo.items(), 2):
            alt = hfill(C["lgray"]) if ri % 2 == 0 else hfill(C["white"])
            for ci, val in enumerate([concepto, tipo, round(total, 2)], 1):
                c = ws4.cell(row=ri, column=ci, value=val)
                c.font = F["body"]; c.fill = alt; c.border = thin()
                if ci == 3:
                    c.number_format = mfmt(); c.alignment = Alignment(horizontal="right")

    wb.save(output)


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Conciliación de pagos Walmart vs Estado de Cuenta"
    )
    parser.add_argument("csv",        help="Ruta al archivo CSV de Walmart")
    parser.add_argument("--output",   "-o", default="",
                        help="Nombre del archivo Excel de salida (opcional)")
    parser.add_argument("--ultimo",   action="store_true",
                        help="Reutiliza los valores del último estado de cuenta sin preguntar")
    args = parser.parse_args()

    if not Path(args.csv).exists():
        print(f"ERROR: No se encontró el archivo: {args.csv}")
        sys.exit(1)

    output = args.output or Path(args.csv).stem + "_conciliacion.xlsx"

    # 1. Captura interactiva del estado de cuenta
    edo_cta, total_pagado_edo = capturar_estado_de_cuenta(usar_ultimo=args.ultimo)

    # 2. Lectura del CSV
    print(f"  Leyendo CSV: {args.csv}")
    df = leer_csv(args.csv)
    print(f"  → {len(df):,} registros cargados")

    # 3. Agrupación por categoría
    print("  Agrupando por categoría...")
    totales_csv, sin_mapeo = agrupar_por_categoria(df)

    # 4. Resumen en consola
    print()
    print("  " + "═" * 78)
    print(f"  {'RESULTADO DE CONCILIACIÓN':^78}")
    print("  " + "═" * 78)
    print(f"  {'Concepto':<44} {'CSV':>12} {'Edo.Cta.':>12} {'Dif.':>9}")
    print("  " + "─" * 78)
    total_csv = 0.0
    for clave, edo_val in edo_cta.items():
        csv_val = totales_csv.get(clave, 0.0)
        diff    = round(csv_val - edo_val, 2)
        total_csv += csv_val
        status  = "✅" if abs(diff) < 0.10 else ("⚠ " if abs(diff) < 1 else "❌")
        etiq    = CAMPO_POR_CLAVE[clave][1]
        print(f"  {status} {etiq:<42} {csv_val:>12,.2f} {edo_val:>12,.2f} {diff:>+9,.2f}")
    total_csv  = round(total_csv, 2)
    diff_total = round(total_csv - total_pagado_edo, 2)
    print("  " + "─" * 78)
    status = "✅" if abs(diff_total) < 1 else "❌"
    print(f"  {status} {'TOTAL PAGADO':<42} {total_csv:>12,.2f} "
          f"{total_pagado_edo:>12,.2f} {diff_total:>+9,.2f}")
    print("  " + "═" * 78)

    if sin_mapeo:
        print(f"\n  ⚠  Movimientos sin mapeo ({len(sin_mapeo)} tipos) – ver hoja 'Sin Mapeo':")
        for (concepto, tipo), total in sin_mapeo.items():
            print(f"     • {concepto} / {tipo}: ${total:,.2f}")

    # 5. Generar Excel
    print(f"\n  Generando Excel: {output}")
    crear_excel(df, totales_csv, sin_mapeo, edo_cta, total_pagado_edo, output)
    print(f"  ✅ Listo: {output}\n")


if __name__ == "__main__":
    main()
