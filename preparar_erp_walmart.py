"""
preparar_erp_walmart.py
=======================
Prepara el CSV de pagos de Walmart para su conciliación contra el ERP.

Lógica de estandarización de identificadores:
  • Walmart Fulfillment  →  FULL-{order_id}
  • Seller Fulfillment   →  Purchase Order# obtenido vía API Walmart MX

Uso:
    python preparar_erp_walmart.py <archivo_csv> [--output <archivo.xlsx>] [--cache-po]

Opciones:
    --cache-po    Reutiliza PO# ya obtenidos en ejecuciones anteriores.
                  Solo llama a la API para órdenes nuevas.

Ejemplos:
    python preparar_erp_walmart.py Atlas_03-2026.csv
    python preparar_erp_walmart.py Atlas_03-2026.csv --cache-po
    python preparar_erp_walmart.py Atlas_03-2026.csv --output ERP_Marzo2026.xlsx

Variables de entorno requeridas (.env en la misma carpeta del script):
    WALMART_CLIENT_ID=394264e4-d823-413d-a2d9-4eed794516ce
    WALMART_CLIENT_SECRET=NvTi9Uxizp...

Dependencias:
    pip install pandas openpyxl requests python-dotenv
"""

import sys
import json
import time
import uuid
import argparse
import csv
import base64
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd
    import requests
    from dotenv import load_dotenv
    import os
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Instala dependencias con:")
    print("  pip install pandas openpyxl requests python-dotenv")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════════════════════

load_dotenv(Path(__file__).parent / ".env", override=True)

WALMART_CLIENT_ID     = os.getenv("WALMART_CLIENT_ID", "")
WALMART_CLIENT_SECRET = os.getenv("WALMART_CLIENT_SECRET", "")

# Endpoints exactos Walmart MX
TOKEN_URL       = "https://marketplace.walmartapis.com/v3/token"
ORDERS_CURSOR   = "https://marketplace.walmartapis.com/v3/orders/cursor"

# Caché local de PO# (evita llamadas repetidas a la API)
PO_CACHE_FILE = Path(__file__).parent / ".walmart_po_cache.json"


def get_credenciales_walmart(cuenta: str) -> tuple[str, str]:
    """Retorna (client_id, client_secret) según la cuenta indicada."""
    cuenta = cuenta.lower().strip()
    if cuenta == "atlas" or cuenta == "":
        return (os.getenv("WALMART_CLIENT_ID", ""),
                os.getenv("WALMART_CLIENT_SECRET", ""))
    sufijo = cuenta.upper()
    cid  = os.getenv(f"WALMART_CLIENT_ID_{sufijo}", "")
    csec = os.getenv(f"WALMART_CLIENT_SECRET_{sufijo}", "")
    if not cid or not csec:
        raise ValueError(
            f"Credenciales para cuenta '{cuenta}' no encontradas en .env.\n"
            f"Agrega WALMART_CLIENT_ID_{sufijo} y WALMART_CLIENT_SECRET_{sufijo}."
        )
    return cid, csec


def get_po_cache_file(cuenta: str) -> Path:
    """Ruta del caché PO# según cuenta (atlas usa el archivo por default)."""
    cuenta = cuenta.lower().strip()
    if cuenta in ("atlas", ""):
        return Path(__file__).parent / ".walmart_po_cache.json"
    return Path(__file__).parent / f".walmart_po_cache_{cuenta}.json"


def get_erp_cache_file(cuenta: str) -> Path:
    """Ruta del caché ERP según cuenta."""
    cuenta = cuenta.lower().strip()
    if cuenta in ("atlas", ""):
        return Path(__file__).parent / ".erp_cache.json"
    return Path(__file__).parent / f".erp_cache_{cuenta}.json"

MAX_RETRIES   = 3
RETRY_DELAY   = 2    # segundos entre reintentos
REQUEST_PAUSE = 0.3  # pausa entre llamadas (throttling)

# ── ERP Atlas del Descanso ────────────────────────────────────────────────────
ERP_BASE_URL   = os.getenv("ERP_BASE_URL",  "https://api.atlasdeldescanso.com")
ERP_USERNAME   = os.getenv("ERP_USERNAME",  "")
ERP_PASSWORD   = os.getenv("ERP_PASSWORD",  "")
ERP_CACHE_FILE = Path(__file__).parent / ".erp_cache.json"


# ══════════════════════════════════════════════════════════════════════════════
#  CLIENTE API WALMART MX
# ══════════════════════════════════════════════════════════════════════════════

class WalmartAPIClient:
    """
    Cliente OAuth2 para la API de Walmart México.
    Replica exactamente los headers del curl proporcionado.
    """

    def __init__(self, client_id: str, client_secret: str):
        if not client_id or not client_secret:
            raise ValueError(
                "Credenciales no configuradas.\n"
                "Crea un archivo .env en la carpeta del script:\n"
                "  WALMART_CLIENT_ID=tu_client_id\n"
                "  WALMART_CLIENT_SECRET=tu_client_secret"
            )
        self.client_id     = client_id
        self.client_secret = client_secret
        self._token        = None
        self._token_expiry = 0

    def _basic_auth(self) -> str:
        """Genera el header Authorization: Basic {base64(client_id:client_secret)}."""
        raw = f"{self.client_id}:{self.client_secret}"
        return "Basic " + base64.b64encode(raw.encode()).decode()

    def _correlation_id(self) -> str:
        """Genera un UUID v4 para WM_QOS.CORRELATION_ID."""
        return str(uuid.uuid4())

    def _headers_token(self) -> dict:
        """Headers para el endpoint de token (igual al curl de autenticación)."""
        return {
            "Authorization":           self._basic_auth(),
            "Content-Type":            "application/x-www-form-urlencoded",
            "WM_MARKET":               "mx",
            "WM_QOS.CORRELATION_ID":   self._correlation_id(),
            "WM_SVC.NAME":             "Walmart Marketplace",
            "Accept":                  "application/json",
        }

    def _headers_api(self) -> dict:
        """Headers para llamadas a la API (igual al curl de órdenes)."""
        return {
            "Authorization":           self._basic_auth(),
            "WM_SEC.ACCESS_TOKEN":     self._token_vigente(),
            "WM_MARKET":               "mx",
            "WM_QOS.CORRELATION_ID":   self._correlation_id(),
            "WM_SVC.NAME":             "Walmart Marketplace",
            "Accept":                  "application/json",
        }

    def obtener_token(self) -> str:
        """Solicita un nuevo token OAuth2."""
        resp = requests.post(
            TOKEN_URL,
            headers=self._headers_token(),
            data={"grant_type": "client_credentials"},
            timeout=15
        )
        resp.raise_for_status()
        data = resp.json()
        self._token        = data["access_token"]
        self._token_expiry = time.time() + data.get("expires_in", 900) - 60
        return self._token

    def _token_vigente(self) -> str:
        """Retorna el token activo, renovándolo si expiró."""
        if not self._token or time.time() >= self._token_expiry:
            self.obtener_token()
        return self._token

    def descargar_po_map(self, dias: int = 120) -> dict:
        """
        Descarga todas las órdenes del período via cursor paginado y retorna
        un dict { customerOrderId: purchaseOrderId } en memoria (sin caché en disco).

        Endpoint:
            GET /v3/orders/cursor
                ?createdStartDate=NOW-{dias}DAYS
                &createdEndDate=NOW
                &limit=100
                &cursorMark=*
                &statusCodeFilter=any

        ~3-5 llamadas API para un mes típico (vs N llamadas uno a uno).
        """
        cursor    = "*"
        pagina    = 1
        po_map    = {}
        fecha_ini = f"NOW-{dias}DAYS"

        print(f"\n  Descargando órdenes ({fecha_ini} → NOW) para mapeo PO#...")

        while True:
            params = {
                "createdStartDate": fecha_ini,
                "createdEndDate":   "NOW",
                "limit":            100,
                "cursorMark":       cursor,
                "statusCodeFilter": "any",
            }

            resp = None
            for intento in range(1, MAX_RETRIES + 1):
                try:
                    resp = requests.get(
                        ORDERS_CURSOR,
                        headers=self._headers_api(),
                        params=params,
                        timeout=20
                    )
                    if resp.status_code == 200:
                        break
                    elif resp.status_code == 401:
                        self._token = None
                        print(f"      🔄 Token expirado, renovando...")
                    elif resp.status_code == 429:
                        wait = int(resp.headers.get("Retry-After", RETRY_DELAY * intento))
                        print(f"      ⏳ Rate limit. Esperando {wait}s...")
                        time.sleep(wait)
                    else:
                        print(f"      ⚠  HTTP {resp.status_code} (intento {intento}/{MAX_RETRIES})")
                        time.sleep(RETRY_DELAY * intento)
                except requests.exceptions.Timeout:
                    print(f"      ⏱  Timeout (intento {intento}/{MAX_RETRIES})")
                    time.sleep(RETRY_DELAY * intento)
                except requests.exceptions.RequestException as e:
                    print(f"      ❌ Error de red: {e}")
                    time.sleep(RETRY_DELAY * intento)
            else:
                print(f"  ⚠  Página {pagina} falló tras {MAX_RETRIES} intentos, abortando descarga.")
                break

            data        = resp.json()
            ordenes     = data.get("order", [])
            total_meta  = data.get("meta", {}).get("totalCount", "?")
            next_cursor = data.get("meta", {}).get("nextCursorMark")

            for orden in ordenes:
                cid = orden.get("customerOrderId")
                pid = orden.get("purchaseOrderId")
                if cid and pid:
                    po_map[cid] = pid

            print(f"  Página {pagina:>3} → {len(ordenes):>3} órdenes  "
                  f"(acumulado: {len(po_map):>5} / {total_meta})")

            if not next_cursor or next_cursor == cursor or next_cursor == "-1":
                break

            cursor  = next_cursor
            pagina += 1
            time.sleep(REQUEST_PAUSE)

        print(f"  ✅ Mapeo PO# completo: {len(po_map):,} órdenes en memoria")
        return po_map


# ══════════════════════════════════════════════════════════════════════════════
#  CLIENTE API ERP (Atlas del Descanso)
# ══════════════════════════════════════════════════════════════════════════════

class ERPAPIClient:
    """Cliente JWT para la API interna del ERP Atlas del Descanso."""

    def __init__(self, base_url: str, username: str, password: str):
        if not username or not password:
            raise ValueError(
                "Credenciales ERP no configuradas.\n"
                "Agrega en .env:\n  ERP_USERNAME=admin\n  ERP_PASSWORD=tu_password"
            )
        self.base_url      = base_url.rstrip("/")
        self.username      = username
        self.password      = password
        self._token        = None
        self._token_expiry = 0

    def autenticar(self) -> str:
        resp = requests.post(
            f"{self.base_url}/auth/login",
            json={"username": self.username, "password": self.password},
            timeout=15,
        )
        resp.raise_for_status()
        token = resp.json()["access_token"]
        # Decodifica exp del JWT sin verificar firma
        try:
            payload = token.split(".")[1]
            payload += "=" * (-len(payload) % 4)
            exp = json.loads(base64.b64decode(payload))["exp"]
            self._token_expiry = exp - 60
        except Exception:
            self._token_expiry = time.time() + 3540  # 59 min por default
        self._token = token
        return self._token

    def _token_vigente(self) -> str:
        if not self._token or time.time() >= self._token_expiry:
            self.autenticar()
        return self._token

    def consultar_orden(self, customer_order_id: str) -> list | None:
        """
        GET /ventas/venta-id/{customer_order_id}
        Retorna lista de documentos ERP o None si no se encontró.
        """
        for intento in range(1, MAX_RETRIES + 1):
            try:
                resp = requests.get(
                    f"{self.base_url}/ventas/venta-id/{customer_order_id}",
                    headers={
                        "Authorization": f"Bearer {self._token_vigente()}",
                        "Content-Type": "application/json",
                    },
                    timeout=15,
                )
                if resp.status_code == 200:
                    data = resp.json()
                    return data if isinstance(data, list) else []
                elif resp.status_code in (404, 204):
                    return None
                elif resp.status_code == 401:
                    self._token = None
                    print("      🔄 Token ERP expirado, renovando...")
                elif resp.status_code == 429:
                    wait = int(resp.headers.get("Retry-After", RETRY_DELAY * intento))
                    print(f"      ⏳ Rate limit ERP. Esperando {wait}s...")
                    time.sleep(wait)
                else:
                    print(f"      ⚠  ERP HTTP {resp.status_code} (intento {intento}/{MAX_RETRIES})")
                    time.sleep(RETRY_DELAY * intento)
            except requests.exceptions.Timeout:
                print(f"      ⏱  Timeout ERP (intento {intento}/{MAX_RETRIES})")
                time.sleep(RETRY_DELAY * intento)
            except requests.exceptions.RequestException as e:
                print(f"      ❌ Error de red ERP: {e}")
                time.sleep(RETRY_DELAY * intento)
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  CACHÉ ERP
# ══════════════════════════════════════════════════════════════════════════════

def cargar_erp_cache(cuenta: str = "atlas") -> dict:
    f = get_erp_cache_file(cuenta)
    if f.exists():
        try:
            return json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def guardar_erp_cache(cache: dict, cuenta: str = "atlas"):
    get_erp_cache_file(cuenta).write_text(
        json.dumps(cache, indent=2, ensure_ascii=False), encoding="utf-8"
    )


# ══════════════════════════════════════════════════════════════════════════════
#  ENRIQUECIMIENTO CON DATOS ERP
# ══════════════════════════════════════════════════════════════════════════════

def extraer_datos_erp(registros: list) -> dict:
    """
    Extrae campos clave de la respuesta ERP para una orden.
    Toma la Factura más reciente y el Pedido que le corresponde.
    Detecta DevVentas (devoluciones/notas de crédito).
    """
    if not registros:
        return {"ERP_En_Sistema": False}

    def fecha_sort(r):
        f = r.get("FechaEmision") or ""
        return f[:19]  # ordena por fecha ISO

    facturas  = sorted(
        [r for r in registros if "Factura" in str(r.get("Mov", ""))],
        key=fecha_sort, reverse=True
    )
    pedidos   = sorted(
        [r for r in registros if "Pedido"  in str(r.get("Mov", ""))],
        key=fecha_sort, reverse=True
    )
    devventas = [r for r in registros if "DevVenta" in str(r.get("Mov", ""))]

    # Factura más reciente (si existe)
    factura = facturas[0] if facturas else None

    # Pedido que corresponde a la factura activa (misma fecha preferida)
    if factura and pedidos:
        fac_fecha = fecha_sort(factura)
        pedido = next(
            (p for p in pedidos if fecha_sort(p) == fac_fecha),
            pedidos[0]   # fallback: pedido más reciente
        )
    else:
        pedido = pedidos[0] if pedidos else None

    # Fuente de montos: Factura si existe, si no el Pedido más reciente
    fuente = factura if factura else pedido

    dev_ids = ", ".join(r["MovID"].strip() for r in devventas) if devventas else ""

    return {
        "ERP_En_Sistema":       True,
        "ERP_Factura":          factura["MovID"].strip()     if factura else "",
        "ERP_Factura_Estatus":  factura["Estatus"].strip()   if factura else "",
        "ERP_PrecioTotal":      fuente["PrecioTotal"]        if fuente  else None,
        "ERP_Importe":          fuente["Importe"]            if fuente  else None,
        "ERP_Impuestos":        fuente["Impuestos"]          if fuente  else None,
        "ERP_FechaEmision":     fuente["FechaEmision"]       if fuente  else "",
        "ERP_Pedido":           pedido["MovID"].strip()      if pedido  else "",
        "ERP_Pedido_Estatus":   pedido["Estatus"].strip()    if pedido  else "",
        "ERP_DevVenta":         dev_ids,
    }


def enriquecer_con_erp(df: pd.DataFrame, cuenta: str = "atlas") -> pd.DataFrame:
    """
    Consulta el ERP por cada orden usando el ERP_ID ya resuelto por preparar_para_erp
    y agrega columnas ERP_* al DataFrame.
      - Seller Fulfillment → ERP_ID = PO# (purchaseOrderId)
      - Walmart Fulfillment → ERP_ID = FULL-{customerOrderId}
      - Cargos de plataforma → ERP_ID vacío, se omiten
    """
    cols_erp = [
        "ERP_En_Sistema", "ERP_Factura", "ERP_Factura_Estatus",
        "ERP_PrecioTotal", "ERP_Importe", "ERP_Impuestos", "ERP_FechaEmision",
        "ERP_Pedido", "ERP_Pedido_Estatus", "ERP_DevVenta",
    ]
    for col in cols_erp:
        df[col] = None

    df["_oid_erp"] = df["Orderline Number"].apply(extraer_order_id)
    order_ids = df["_oid_erp"].unique().tolist()

    # Construye erp_query_id desde la columna ERP_ID (ya resuelta por preparar_para_erp)
    oid_to_erp_id = (
        df.drop_duplicates("_oid_erp")
          .set_index("_oid_erp")["ERP_ID"]
          .to_dict()
    )

    erp_query_id = {}
    ids_sin_erp  = []
    for oid in order_ids:
        erp_id = oid_to_erp_id.get(oid, "")
        if erp_id:
            erp_query_id[oid] = erp_id
        else:
            ids_sin_erp.append(oid)   # cargos de plataforma (SEM, etc.)

    if ids_sin_erp:
        print(f"\n  ERP: {len(ids_sin_erp)} cargo(s) de plataforma sin ERP_ID → omitidos")

    total      = len(erp_query_id)
    seller_cnt = sum(1 for q in erp_query_id.values() if not q.startswith("FULL-"))
    wfs_cnt    = sum(1 for q in erp_query_id.values() if q.startswith("FULL-"))
    print(f"\n  ERP API: {total} órdenes ({seller_cnt} Seller por PO# + {wfs_cnt} WFS por FULL-id)")

    if not ERP_USERNAME or not ERP_PASSWORD:
        print("\n  ❌ Credenciales ERP no configuradas. Agrega en .env:")
        print("       ERP_USERNAME=admin")
        print("       ERP_PASSWORD=tu_password")
        print("  → Continuando sin datos ERP.\n")
        df.drop(columns=["_oid_erp"], inplace=True)
        return df

    try:
        cliente_erp = ERPAPIClient(ERP_BASE_URL, ERP_USERNAME, ERP_PASSWORD)
        cliente_erp.autenticar()
        print("  ✅ Autenticación ERP exitosa\n")
    except Exception as e:
        print(f"  ❌ Error de autenticación ERP: {e}")
        df.drop(columns=["_oid_erp"], inplace=True)
        return df

    cache = {}
    ancho = len(str(total))
    for i, (cid, qid) in enumerate(erp_query_id.items(), 1):
        print(f"  [{i:>{ancho}}/{total}]  {qid}...", end=" ", flush=True)
        registros = cliente_erp.consultar_orden(qid)
        if registros:
            cache[cid] = registros
            d = extraer_datos_erp(registros)
            print(f"✅  Factura {d.get('ERP_Factura','—')}  Pedido {d.get('ERP_Pedido','—')}")
        else:
            cache[cid] = []
            print("⚠   No encontrado en ERP")
        time.sleep(REQUEST_PAUSE)

    # Aplica datos ERP al DataFrame
    for oid, registros in cache.items():
        mask  = df["_oid_erp"] == oid
        datos = extraer_datos_erp(registros)
        for col, val in datos.items():
            df.loc[mask, col] = val

    # ERP_Diferencia por orden: suma de Venta/Ingreso vs PrecioTotal del ERP
    df["ERP_PrecioTotal"] = pd.to_numeric(df["ERP_PrecioTotal"], errors="coerce")
    df["ERP_Importe"]     = pd.to_numeric(df["ERP_Importe"],     errors="coerce")
    df["ERP_Impuestos"]   = pd.to_numeric(df["ERP_Impuestos"],   errors="coerce")
    mask_venta = (df["Concept"] == "Venta") & (df["Income / Outcome"] == "Ingreso")
    totales_wm = df[mask_venta].groupby("_oid_erp")["Amount"].sum()
    df["ERP_Diferencia"] = (
        df["_oid_erp"].map(totales_wm) - df["ERP_PrecioTotal"]
    ).round(2)
    df.loc[df["ERP_PrecioTotal"].isna(), "ERP_Diferencia"] = None

    en   = (df["ERP_En_Sistema"] == True).sum()
    no   = (df["ERP_En_Sistema"] == False).sum()
    nulo = df["ERP_En_Sistema"].isna().sum()
    print(f"\n  ERP: {en} con datos  |  {no} no encontrados  |  {nulo} sin consultar")

    df.drop(columns=["_oid_erp"], inplace=True)
    return df


# ══════════════════════════════════════════════════════════════════════════════
#  LECTURA DEL CSV
# ══════════════════════════════════════════════════════════════════════════════

def leer_csv(ruta: str) -> pd.DataFrame:
    """Lee el CSV de Walmart (2 filas de encabezado: inglés + español)."""
    with open(ruta, encoding="utf-8") as f:
        lines = f.readlines()
    eng_header = next(csv.reader([lines[0]]))
    rows = []
    for line in lines[2:]:
        if not line.strip():
            continue
        row = next(csv.reader([line]))
        record = {
            eng_header[i]: row[i] if i < len(row) else ""
            for i in range(len(eng_header))
        }
        rows.append(record)
    df = pd.DataFrame(rows)
    df["Amount"]           = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)
    df["Concept"]          = df["Concept"].str.strip()
    df["Income / Outcome"] = df["Income / Outcome"].str.strip()
    df["FulfilledBy"]      = df["FulfilledBy"].str.strip()
    df["Orderline Number"] = df["Orderline Number"].str.strip()
    return df


# ══════════════════════════════════════════════════════════════════════════════
#  TRANSFORMACIÓN DE IDENTIFICADORES
# ══════════════════════════════════════════════════════════════════════════════

def extraer_order_id(orderline_number: str) -> str:
    """
    Extrae el customerOrderId del Orderline Number.
    '600000076000151-3'   → '600000076000151'
    '600000038614208-2-1' → '600000038614208'  (formato con doble sufijo)
    '11516562_SEM'        → '11516562_SEM'     (cargos de plataforma, sin guion)
    """
    if "-" in orderline_number:
        return orderline_number.split("-", 1)[0]   # parte desde la izquierda
    return orderline_number


def construir_erp_id_wfs(order_id: str) -> str:
    return f"FULL-{order_id}"


# ══════════════════════════════════════════════════════════════════════════════
#  PREPARACIÓN DEL DATAFRAME PARA ERP
# ══════════════════════════════════════════════════════════════════════════════

def preparar_para_erp(df: pd.DataFrame, cuenta: str = "atlas") -> pd.DataFrame:
    """
    Agrega columnas ERP_ID y ERP_Tipo al DataFrame según FulfilledBy:
      - Walmart Fulfillment → FULL-{order_id}   (sin API)
      - Seller Fulfillment  → PO# via batch API  (descarga todas las órdenes del período)
      - Sin FulfilledBy     → ERP_Pendientes (ej: cargos SEM)
    """
    df = df.copy()
    df["_order_id"] = df["Orderline Number"].apply(extraer_order_id)
    df["ERP_ID"]    = ""
    df["ERP_Tipo"]  = ""

    mask_wfs    = df["FulfilledBy"].str.lower().str.contains("walmart", na=False)
    mask_seller = df["FulfilledBy"].str.lower().str.contains("seller",  na=False)
    mask_otros  = ~mask_wfs & ~mask_seller

    # ── Walmart Fulfillment: no requiere API ──────────────────────────────────
    df.loc[mask_wfs, "ERP_ID"]   = df.loc[mask_wfs, "_order_id"].apply(construir_erp_id_wfs)
    df.loc[mask_wfs, "ERP_Tipo"] = "Walmart Fulfillment"

    # ── Sin FulfilledBy (cargos de plataforma: SEM, etc.) ────────────────────
    df.loc[mask_otros, "ERP_Tipo"] = "Sin Fulfillment (plataforma)"

    # ── Seller Fulfillment ────────────────────────────────────────────────────
    df.loc[mask_seller, "ERP_Tipo"] = "Seller Fulfillment"

    seller_order_ids = df.loc[mask_seller, "_order_id"].unique()
    total_seller     = len(seller_order_ids)

    if total_seller == 0:
        df.drop(columns=["_order_id"], inplace=True)
        return df

    # ── BATCH: descarga todas las órdenes del período en memoria ─────────────
    print(f"\n  Seller Fulfillment: {total_seller} órdenes únicas → descarga batch PO#...")

    try:
        cid, csec = get_credenciales_walmart(cuenta)
    except ValueError as e:
        print(f"\n  ❌ {e}")
        sys.exit(1)

    if not cid or not csec:
        print("\n  ❌ Credenciales no configuradas. Crea un archivo .env con:")
        print("       WALMART_CLIENT_ID=tu_client_id")
        print("       WALMART_CLIENT_SECRET=tu_client_secret")
        sys.exit(1)

    try:
        cliente = WalmartAPIClient(cid, csec)
        cliente.obtener_token()
        print("  ✅ Token obtenido")
    except Exception as e:
        print(f"  ❌ Error de autenticación: {e}")
        sys.exit(1)

    po_map = cliente.descargar_po_map(dias=120)

    df.loc[mask_seller, "ERP_ID"] = (
        df.loc[mask_seller, "_order_id"].map(po_map).fillna("")
    )

    resueltos  = df.loc[mask_seller, "ERP_ID"].ne("").sum()
    pendientes = df.loc[mask_seller, "ERP_ID"].eq("").sum()
    print(f"\n  Resultado: {resueltos} con PO#  |  {pendientes} pendientes")

    df.drop(columns=["_order_id"], inplace=True)
    return df


# ══════════════════════════════════════════════════════════════════════════════
#  ESTILOS EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def hfill(h):  return PatternFill("solid", start_color=h, fgColor=h)
def thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)
def mfmt():    return '#,##0.00;(#,##0.00);"-"'

C = dict(blue="2E75B6", blue2="0070C0", lgray="F2F2F2", white="FFFFFF",
         yellow="FFF2CC", secbg="D6E4F0", green_bg="E2EFDA", blue_bg="DEEAF1",
         red_bg="FFE0E0")
F = {
    "title": Font(name="Arial", bold=True, size=13, color="FFFFFF"),
    "hdr":   Font(name="Arial", bold=True, size=10, color="FFFFFF"),
    "body":  Font(name="Arial", size=10),
    "bold":  Font(name="Arial", bold=True, size=10),
    "ital":  Font(name="Arial", size=9,  italic=True, color="595959"),
    "grn":   Font(name="Arial", size=10, bold=True, color="006100"),
    "blue":  Font(name="Arial", size=10, bold=True, color="0070C0"),
    "red":   Font(name="Arial", size=10, color="CC0000"),
    "org":   Font(name="Arial", size=10, color="9C5700"),
}

# Anchos de columna por nombre
COL_ANCHOS = {
    # ── Walmart ERP ──────────────────────────────────────────────────────────
    "ERP_ID": 26, "ERP_Tipo": 24,
    # ── ERP Atlas ────────────────────────────────────────────────────────────
    "ERP_En_Sistema": 14, "ERP_Factura": 16, "ERP_Factura_Estatus": 16,
    "ERP_PrecioTotal": 16, "ERP_Importe": 16, "ERP_Impuestos": 14,
    "ERP_FechaEmision": 18, "ERP_Pedido": 16, "ERP_Pedido_Estatus": 16,
    "ERP_Diferencia": 16, "ERP_DevVenta": 20,
    # ── CSV ──────────────────────────────────────────────────────────────────
    "Payment Date": 14, "PartnerId": 14, "VendorNumber": 16,
    "Store": 30, "Orderline Number": 28, "Amount": 16,
    "Income / Outcome": 16, "Concept": 24, "Date of Retention/Refund": 22,
    "Invoice Number": 18, "Invoice Date": 14, "Credit Note": 14,
    "Date for CN": 14, "FulfilledBy": 22, "Fulfillment LineId": 20,
    "Shipment Number": 20, "Tracking Number": 28, "Return Number": 18,
    "Quantity": 10, "Total Walmart Funded Savings Program": 20,
}


def _set_col_widths(ws, cols):
    for i, col in enumerate(cols, 1):
        letra = chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
        ws.column_dimensions[letra].width = COL_ANCHOS.get(col, 16)


def escribir_hoja_detalle(wb, df_hoja, titulo_hoja, titulo_banner, color_banner):
    if df_hoja.empty:
        return

    ws = wb.create_sheet(titulo_hoja)

    # Orden de columnas: Walmart ERP → Atlas ERP → resto
    cols_wm_erp  = ["ERP_ID", "ERP_Tipo"]
    cols_atl_erp = [
        "ERP_En_Sistema", "ERP_Factura", "ERP_Factura_Estatus",
        "ERP_Pedido", "ERP_Pedido_Estatus",
        "ERP_PrecioTotal", "ERP_Importe", "ERP_Impuestos",
        "ERP_FechaEmision", "ERP_Diferencia",
    ]
    cols_todas   = cols_wm_erp + cols_atl_erp
    cols_orig    = [c for c in df_hoja.columns if c not in cols_todas]
    cols         = cols_wm_erp + [c for c in cols_atl_erp if c in df_hoja.columns] + cols_orig
    _set_col_widths(ws, cols)

    # Banner
    ultima_col = chr(64 + len(cols)) if len(cols) <= 26 else "A" + chr(64 + len(cols) - 26)
    ws.merge_cells(f"A1:{ultima_col}1")
    ws["A1"] = titulo_banner
    ws["A1"].font = F["title"]; ws["A1"].fill = hfill(color_banner)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Info
    ws["A2"] = "Generado:"; ws["B2"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws["C2"] = "Registros:"; ws["D2"] = len(df_hoja)
    for cel in ["A2", "B2", "C2", "D2"]:
        ws[cel].fill = hfill(C["lgray"]); ws[cel].font = F["body"]
    ws["A2"].font = F["bold"]; ws["C2"].font = F["bold"]
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    # Encabezados
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=4, column=ci, value=col)
        c.font = F["hdr"]; c.fill = hfill(C["blue"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin()
    ws.row_dimensions[4].height = 28

    # Datos
    for ri, row_data in enumerate(df_hoja[cols].itertuples(index=False), 5):
        alt = hfill(C["lgray"]) if ri % 2 == 0 else hfill(C["white"])

        for ci, (col, val) in enumerate(zip(cols, row_data), 1):
            val_str = "" if (val is None or (isinstance(val, float) and pd.isna(val))) else val
            c = ws.cell(row=ri, column=ci, value=val_str)
            c.font  = F["body"]
            c.fill  = alt
            c.border = thin()
            c.alignment = Alignment(vertical="center")

            if col == "Amount":
                c.number_format = mfmt()
                c.alignment = Alignment(horizontal="right", vertical="center")

            # Color ERP_ID según tipo
            if col == "ERP_ID":
                if str(val_str).startswith("FULL-"):
                    c.font = F["grn"]; c.fill = hfill(C["green_bg"])
                elif val_str and str(val_str) != "":
                    c.font = F["blue"]; c.fill = hfill(C["blue_bg"])
                else:
                    c.font = F["red"]; c.fill = hfill(C["red_bg"])

            elif col == "ERP_Tipo":
                if "Walmart" in str(val_str):
                    c.font = Font(name="Arial", size=10, color="006100")
                elif "Seller" in str(val_str):
                    c.font = Font(name="Arial", size=10, color="0070C0")
                else:
                    c.font = F["org"]

            elif col == "ERP_En_Sistema":
                if val_str is True or val_str == True:
                    c.font = F["grn"]; c.fill = hfill(C["green_bg"]); c.value = "✅ Sí"
                elif val_str is False or val_str == False:
                    c.font = F["red"]; c.fill = hfill(C["red_bg"]); c.value = "❌ No"
                else:
                    c.fill = alt

            elif col in ("ERP_PrecioTotal", "ERP_Importe", "ERP_Impuestos"):
                c.number_format = mfmt()
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.fill = alt

            elif col == "ERP_Diferencia":
                c.number_format = mfmt()
                c.alignment = Alignment(horizontal="right", vertical="center")
                try:
                    v = float(val_str) if val_str not in ("", None) else None
                    if v is None:
                        c.fill = alt
                    elif abs(v) < 0.10:
                        c.font = F["grn"]; c.fill = hfill(C["green_bg"])
                    elif abs(v) < 10:
                        c.font = F["org"]; c.fill = alt
                    else:
                        c.font = F["red"]; c.fill = hfill(C["red_bg"])
                except (TypeError, ValueError):
                    c.fill = alt

            elif col in ("ERP_Factura_Estatus", "ERP_Pedido_Estatus"):
                estatus = str(val_str).strip()
                if estatus == "CONCLUIDO":
                    c.font = F["grn"]; c.fill = hfill(C["green_bg"])
                elif estatus and estatus not in ("", "None"):
                    c.font = F["org"]; c.fill = alt
                else:
                    c.fill = alt

        ws.row_dimensions[ri].height = 16


# ══════════════════════════════════════════════════════════════════════════════
#  GENERACIÓN DEL EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def crear_excel(df: pd.DataFrame, ruta_csv: str, output: str):
    wb = Workbook()
    ws_tmp = wb.active; ws_tmp.title = "_tmp"

    mask_wfs    = df["FulfilledBy"].str.lower().str.contains("walmart", na=False)
    mask_seller = df["FulfilledBy"].str.lower().str.contains("seller",  na=False)
    mask_otros  = ~mask_wfs & ~mask_seller

    df_wfs      = df[mask_wfs].copy()
    df_seller   = df[mask_seller].copy()
    df_otros    = df[mask_otros].copy()

    df_seller_ok   = df_seller[df_seller["ERP_ID"].ne("") & df_seller["ERP_ID"].notna()].copy()
    df_pendientes  = df_seller[df_seller["ERP_ID"].eq("") | df_seller["ERP_ID"].isna()].copy()
    # Registros sin FulfilledBy también van a pendientes
    df_pendientes  = pd.concat([df_pendientes, df_otros], ignore_index=True)

    # ── Hoja 1: Resumen ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Resumen")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20

    ws.merge_cells("A1:C1")
    ws["A1"] = "PREPARACIÓN PARA CONCILIACIÓN ERP – WALMART"
    ws["A1"].font = F["title"]; ws["A1"].fill = hfill(C["blue2"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    meta = [
        ("Archivo CSV:",     Path(ruta_csv).name),
        ("Fecha de pago:",   df["Payment Date"].dropna().iloc[0] if not df.empty else "N/D"),
        ("Generado:",        datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Total registros:", len(df)),
    ]
    for r, (lbl, val) in enumerate(meta, 2):
        ws.cell(row=r, column=1, value=lbl).font = F["bold"]
        ws.cell(row=r, column=2, value=val).font  = F["body"]
        for ci in range(1, 4):
            ws.cell(row=r, column=ci).fill   = hfill(C["lgray"])
            ws.cell(row=r, column=ci).border = thin()
        ws.row_dimensions[r].height = 16

    r = len(meta) + 3
    ws.merge_cells(f"A{r}:C{r}")
    ws.cell(row=r, column=1, value="DESGLOSE POR TIPO DE FULFILLMENT")
    ws.cell(row=r, column=1).font = F["bold"]
    ws.cell(row=r, column=1).fill = hfill(C["secbg"])
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(row=r, column=1).border = thin()
    ws.row_dimensions[r].height = 18; r += 1

    for ci, h in enumerate(["Tipo", "Registros", "Monto Total ($)"], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = F["hdr"]; c.fill = hfill(C["blue"])
        c.alignment = Alignment(horizontal="center"); c.border = thin()
    ws.row_dimensions[r].height = 20; r += 1

    filas_resumen = [
        ("Walmart Fulfillment (FULL-*)",         len(df_wfs),        df_wfs["Amount"].sum(),       C["green_bg"], "006100"),
        ("Seller Fulfillment – con PO#",          len(df_seller_ok),  df_seller_ok["Amount"].sum(), C["blue_bg"],  "0070C0"),
        ("Pendientes sin PO# / sin FulfilledBy",  len(df_pendientes), df_pendientes["Amount"].sum(),C["red_bg"],   "CC0000"),
        ("TOTAL",                                 len(df),            df["Amount"].sum(),            C["yellow"],   "000000"),
    ]
    for label, cnt, monto, bg, fc in filas_resumen:
        es_total = label == "TOTAL"
        ws.cell(row=r, column=1, value=label).font  = Font(name="Arial", size=10, bold=es_total, color=fc)
        ws.cell(row=r, column=2, value=cnt).font    = Font(name="Arial", size=10, bold=es_total, color=fc)
        ws.cell(row=r, column=3, value=monto).font  = Font(name="Arial", size=10, bold=es_total, color=fc)
        ws.cell(row=r, column=3).number_format = mfmt()
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
        for ci in range(1, 4):
            ws.cell(row=r, column=ci).fill   = hfill(bg)
            ws.cell(row=r, column=ci).border = thin()
        ws.row_dimensions[r].height = 18; r += 1

    r += 1
    ws.merge_cells(f"A{r}:C{r}")
    ws.cell(row=r, column=1,
            value=("FORMATO ERP_ID:  "
                   "Walmart Fulfillment → FULL-{order_id}   |   "
                   "Seller Fulfillment → PO# obtenido vía API Walmart MX  "
                   "(GET /v3/orders/cursor?customerOrderId=...)"))
    ws.cell(row=r, column=1).font  = F["ital"]
    ws.cell(row=r, column=1).fill  = hfill(C["yellow"])
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, horizontal="left", indent=1)
    ws.cell(row=r, column=1).border = thin()
    ws.row_dimensions[r].height = 32

    # ── Hojas de detalle ──────────────────────────────────────────────────────
    escribir_hoja_detalle(wb, df_wfs,
        "WFS_Detalle",
        "WALMART FULFILLMENT – Listos para ERP  (ERP_ID = FULL-*)",
        "375623")

    escribir_hoja_detalle(wb, df_seller_ok,
        "Seller_Detalle",
        "SELLER FULFILLMENT – Con PO# (listos para ERP)",
        C["blue2"])

    if not df_pendientes.empty:
        escribir_hoja_detalle(wb, df_pendientes,
            "ERP_Pendientes",
            "⚠  PENDIENTES – Sin PO# o sin FulfilledBy  (seguimiento manual)",
            "C00000")

    # ── Hoja: Conciliación ERP ────────────────────────────────────────────────
    if "ERP_En_Sistema" in df.columns:
        _hoja_erp_conciliacion(wb, df)

    del wb["_tmp"]
    wb.save(output)


def _hoja_erp_conciliacion(wb, df: pd.DataFrame, nombre_hoja: str = "ERP_Conciliacion"):
    """Hoja de resumen por orden: desglose de conceptos CSV + datos ERP."""
    ws = wb.create_sheet(nombre_hoja)

    df2 = df.copy()
    df2["_oid"] = df2["Orderline Number"].apply(extraer_order_id)
    has_cuenta  = "Cuenta" in df2.columns

    # ── Columnas dinámicas de conceptos (Ingreso primero, luego Egreso) ────────
    df2["_col_key"] = df2["Concept"] + " / " + df2["Income / Outcome"]
    orden_tipo = df2.drop_duplicates("_col_key").sort_values(
        ["Income / Outcome", "Concept"], ascending=[False, True]
    )["_col_key"].tolist()

    pivot = (
        df2.groupby(["_oid", "_col_key"])["Amount"]
        .sum().unstack(fill_value=0)
        .reindex(columns=orden_tipo, fill_value=0)
    )

    # ── ERP data por orden ─────────────────────────────────────────────────────
    erp_needed = [
        "ERP_PrecioTotal", "ERP_Factura", "ERP_Factura_Estatus",
        "ERP_Pedido", "ERP_Pedido_Estatus", "ERP_FechaEmision",
        "ERP_En_Sistema", "ERP_DevVenta",
    ]
    erp_first = (
        df2.groupby("_oid")[erp_needed].first()
        if all(c in df2.columns for c in erp_needed)
        else pd.DataFrame(index=df2["_oid"].unique())
    )
    if has_cuenta:
        cuenta_map = df2.groupby("_oid")["Cuenta"].first()

    mask_venta  = (df2["Concept"] == "Venta") & (df2["Income / Outcome"] == "Ingreso")
    mask_devol  = (df2["Concept"] == "Venta") & (df2["Income / Outcome"] == "Egreso")
    wm_ventas   = df2[mask_venta].groupby("_oid")["Amount"].sum()
    wm_devol    = df2[mask_devol].groupby("_oid")["Amount"].sum().abs()
    neto_csv    = df2.groupby("_oid")["Amount"].sum()

    # Tipo de operación por orden
    oids_devol  = set(df2.loc[mask_devol, "_oid"].unique())
    oids_venta  = set(df2.loc[mask_venta, "_oid"].unique())
    def tipo_operacion(oid):
        if oid in oids_devol and oid not in oids_venta:
            return "🔄 Nota de crédito"
        if oid in oids_devol and oid in oids_venta:
            return "⚠ Venta + Devolución"
        return "💰 A pagar proveedor"

    # ── Estructura de columnas ─────────────────────────────────────────────────
    left_cols  = (["Cuenta"] if has_cuenta else []) + ["CustomerOrderID", "Tipo Operación", "Registros"]
    right_cols = [
        "Neto CSV", "ERP_PrecioTotal", "Diferencia",
        "ERP_Factura", "ERP_Factura_Estatus",
        "ERP_Pedido", "ERP_Pedido_Estatus",
        "ERP_FechaEmision", "ERP_DevVenta", "Estado",
    ]
    all_cols  = left_cols + orden_tipo + right_cols
    n_cols    = len(all_cols)

    # Índices (1-based) de columnas numéricas
    conc_start = len(left_cols) + 1
    conc_end   = conc_start + len(orden_tipo) - 1
    neto_ci    = conc_end + 1
    erp_pt_ci  = neto_ci + 1
    diff_ci    = erp_pt_ci + 1
    num_ci     = set(range(conc_start, diff_ci + 1))

    # ── Anchos de columna ──────────────────────────────────────────────────────
    left_widths = {"Cuenta": 10, "CustomerOrderID": 26, "Tipo Operación": 22, "Registros": 10}
    for i, col in enumerate(left_cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = left_widths[col]
    for i in range(conc_start, conc_end + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    right_widths = [14, 16, 14, 16, 18, 16, 18, 20, 22, 14]
    for i, w in enumerate(right_widths, conc_end + 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Banner ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    ws["A1"] = "CONCILIACIÓN ERP – Desglose por Concepto / Tipo vs ERP" + (" (Atlas + Spring)" if has_cuenta else "")
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = hfill(C["blue2"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Encabezados ────────────────────────────────────────────────────────────
    for ci, col in enumerate(all_cols, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = thin()
        if ci <= len(left_cols):
            c.fill = hfill(C["blue"])
        elif ci <= conc_end:
            c.fill = hfill("1F7145") if "Ingreso" in col else hfill("C00000")
        else:
            c.fill = hfill(C["blue2"])
    ws.row_dimensions[2].height = 36

    # ── Filas de datos ─────────────────────────────────────────────────────────
    order_ids = sorted(df2["_oid"].unique())
    r = 3
    for oid in order_ids:
        cnt    = int((df2["_oid"] == oid).sum())
        neto   = round(float(neto_csv.get(oid, 0.0)), 2)
        erp_pt = erp_first.at[oid, "ERP_PrecioTotal"]      if oid in erp_first.index else None
        erp_fac  = erp_first.at[oid, "ERP_Factura"]         if oid in erp_first.index else ""
        erp_fst  = erp_first.at[oid, "ERP_Factura_Estatus"] if oid in erp_first.index else ""
        erp_ped  = erp_first.at[oid, "ERP_Pedido"]          if oid in erp_first.index else ""
        erp_pst  = erp_first.at[oid, "ERP_Pedido_Estatus"]  if oid in erp_first.index else ""
        erp_fech = erp_first.at[oid, "ERP_FechaEmision"]    if oid in erp_first.index else ""
        en_sis   = erp_first.at[oid, "ERP_En_Sistema"]      if oid in erp_first.index else None
        erp_dev  = erp_first.at[oid, "ERP_DevVenta"]        if oid in erp_first.index else ""

        # Para notas de crédito compara Venta/Egreso (valor absoluto) vs ERP
        if oid in oids_devol and oid not in oids_venta:
            wm_v = round(float(wm_devol.get(oid, 0.0)), 2)
        else:
            wm_v = round(float(wm_ventas.get(oid, 0.0)), 2)
        try:
            diff = round(wm_v - float(erp_pt), 2) if erp_pt is not None else None
        except (TypeError, ValueError):
            diff = None

        if en_sis is False:
            estado, bg_row = "❌ No en ERP",       "FFE0E0"
        elif diff is None:
            estado, bg_row = "⚠ Sin datos ERP",   "FFF2CC"
        elif abs(diff) < 0.10:
            estado, bg_row = "✅ Cuadra",           "E2EFDA"
        elif abs(diff) < 10:
            estado, bg_row = f"⚠ Dif ${diff:+,.2f}", "FFF2CC"
        else:
            estado, bg_row = f"❌ Dif ${diff:+,.2f}", "FFE0E0"

        tipo_op   = tipo_operacion(oid)
        conc_vals = [
            round(float(pivot.at[oid, col]), 2) if oid in pivot.index else 0.0
            for col in orden_tipo
        ]
        cuenta_prefix = [cuenta_map.get(oid, "")] if has_cuenta else []
        vals = cuenta_prefix + [oid, tipo_op, cnt] + conc_vals + [
            neto, erp_pt, diff,
            erp_fac, str(erp_fst).strip(),
            erp_ped, str(erp_pst).strip(),
            str(erp_fech)[:10] if erp_fech else "",
            str(erp_dev) if erp_dev else "",
            estado,
        ]

        # Color de fuente para Tipo Operación (col 2 sin cuenta, col 3 con cuenta)
        tipo_color = {
            "💰 A pagar proveedor":  "006100",
            "🔄 Nota de crédito":    "0070C0",
            "⚠ Venta + Devolución": "9C5700",
        }

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill   = hfill(bg_row)
            c.border = thin()
            c.alignment = Alignment(vertical="center")
            tipo_ci = 3 if has_cuenta else 2
            if ci == tipo_ci:
                c.font = Font(name="Arial", size=10, bold=True,
                              color=tipo_color.get(val, "000000"))
            else:
                c.font = Font(name="Arial", size=10)
            if ci in num_ci:
                c.number_format = mfmt()
                c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 16
        r += 1

    # ── Fila totales ───────────────────────────────────────────────────────────
    conc_totals = [
        round(float(pivot[col].sum()), 2) if col in pivot.columns else 0.0
        for col in orden_tipo
    ]
    tot_neto = round(float(neto_csv.sum()), 2)
    tot_erp  = pd.to_numeric(
        df2.groupby("_oid")["ERP_PrecioTotal"].first(), errors="coerce"
    ).sum() if "ERP_PrecioTotal" in df2.columns else None
    tot_wm   = round(float(wm_ventas.sum()), 2)
    tot_diff = round(tot_wm - float(tot_erp), 2) if tot_erp else None

    tot_vals = ["TOTAL", "", len(order_ids)] + conc_totals + [
        tot_neto, tot_erp if tot_erp else None, tot_diff,
        "", "", "", "", "", "", "", "",
    ]
    for ci, val in enumerate(tot_vals, 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.fill   = hfill(C["yellow"])
        c.font   = Font(name="Arial", bold=True, size=10)
        c.border = thin()
        if ci in num_ci:
            c.number_format = mfmt()
            c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[r].height = 18


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Prepara CSV de pagos Walmart para conciliación vs ERP"
    )
    parser.add_argument("csv",        help="Ruta al archivo CSV de Walmart")
    parser.add_argument("--output", "-o", default="",
                        help="Nombre del archivo Excel de salida (opcional)")
    parser.add_argument("--cache-po",  action="store_true",
                        help="Reutiliza caché de PO# previos (solo consulta órdenes nuevas)")
    parser.add_argument("--cache-erp", action="store_true",
                        help="Reutiliza datos ERP ya consultados (evita llamadas al ERP)")
    parser.add_argument("--skip-erp",  action="store_true",
                        help="Omite la consulta al ERP por completo")
    parser.add_argument("--cuenta", default="atlas",
                        help="Cuenta Walmart a procesar: atlas (default) | spring")
    args = parser.parse_args()

    if not Path(args.csv).exists():
        print(f"\n  ERROR: No se encontró el archivo: {args.csv}")
        sys.exit(1)

    output = args.output or Path(args.csv).stem + "_erp.xlsx"

    # 1. Leer CSV
    print(f"\n  Leyendo CSV: {args.csv}")
    df = leer_csv(args.csv)

    mask_wfs    = df["FulfilledBy"].str.lower().str.contains("walmart", na=False)
    mask_seller = df["FulfilledBy"].str.lower().str.contains("seller",  na=False)
    print(f"  → {len(df):,} registros  "
          f"({mask_wfs.sum()} WFS  |  {mask_seller.sum()} Seller  |  "
          f"{(~mask_wfs & ~mask_seller).sum()} sin FulfilledBy)")

    # 2. Preparar ERP_IDs (Walmart)
    df = preparar_para_erp(df, usar_cache=args.cache_po, cuenta=args.cuenta)

    # 3. Enriquecer con datos del ERP Atlas
    if not args.skip_erp:
        df = enriquecer_con_erp(df, usar_cache=args.cache_erp, cuenta=args.cuenta)
    else:
        print("\n  ERP Atlas: omitido (--skip-erp)")

    # 5. Resumen en consola
    df_pend = df[df["ERP_ID"].eq("") | df["ERP_ID"].isna()]
    df_ok   = df[df["ERP_ID"].ne("") & df["ERP_ID"].notna()]

    print()
    print("  " + "═" * 56)
    print(f"  {'RESULTADO':^56}")
    print("  " + "═" * 56)
    print(f"  {'Registros con ERP_ID asignado:':<38} {len(df_ok):>8,}")
    print(f"  {'  └─ FULL-* (Walmart Fulfillment):':<38} {mask_wfs.sum():>8,}")
    print(f"  {'  └─ PO#   (Seller Fulfillment):':<38} {df_ok[df_ok['ERP_Tipo']=='Seller Fulfillment']['ERP_ID'].ne('').sum():>8,}")
    print(f"  {'Registros pendientes (sin ERP_ID):':<38} {len(df_pend):>8,}")
    print("  " + "─" * 56)
    print(f"  {'TOTAL:':<38} {len(df):>8,}")
    print("  " + "═" * 56)

    if len(df_pend) > 0:
        print(f"\n  ⚠  {len(df_pend)} registros sin ERP_ID → hoja 'ERP_Pendientes'")

    # 6. Generar Excel
    print(f"\n  Generando Excel: {output}")
    crear_excel(df, args.csv, output)
    print(f"  ✅ Listo: {output}\n")
    if "ERP_En_Sistema" in df.columns:
        print("  Hojas adicionales:")
        print("    ERP_Conciliacion  – Resumen por orden Walmart vs ERP")


if __name__ == "__main__":
    main()
