"""
run_conciliacion.py
===================
Orquestador del proceso mensual de conciliación Walmart MX.

Produce un Excel unificado con:
  • Detalle ERP clasificado por acción contable
  • Órdenes a pagar proveedor
  • Notas de crédito por devolución
  • Gastos de plataforma
  • Retenciones fiscales
  • Pendientes sin PO#

Uso:
    python run_conciliacion.py <archivo_csv> [opciones]

Opciones:
    --output  -o    Nombre del Excel de salida (default: <csv>_reporte.xlsx)
    --skip-erp      Omite la consulta al ERP por completo
    --cuenta        Cuenta Walmart: atlas (default) | spring

Ejemplo:
    python run_conciliacion.py Atlas_03-2026.csv
    python run_conciliacion.py Spring_03-2026.csv --cuenta spring

Dependencias:
    pip install pandas openpyxl requests python-dotenv
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime

# ── Verifica que los módulos existan en la misma carpeta ─────────────────────
_DIR = Path(__file__).parent
for _mod in ["conciliacion_walmart.py", "preparar_erp_walmart.py"]:
    if not (_DIR / _mod).exists():
        print(f"ERROR: No se encontró {_mod} en {_DIR}")
        sys.exit(1)

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install pandas openpyxl requests python-dotenv")
    sys.exit(1)

# ── Importa funciones de los módulos especializados ──────────────────────────
import importlib.util, types

def _load(name: str):
    spec = importlib.util.spec_from_file_location(name, _DIR / f"{name}.py")
    mod  = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod

_conc = _load("conciliacion_walmart")
_erp  = _load("preparar_erp_walmart")


# ══════════════════════════════════════════════════════════════════════════════
#  CLASIFICACIÓN CONTABLE
#  Lógica de contador ecommerce especializado en Walmart MX
# ══════════════════════════════════════════════════════════════════════════════

# (Concept, Income/Outcome) → acción ERP
ACCION_ERP = {
    # ── A PAGAR PROVEEDOR ────────────────────────────────────────────────────
    # Walmart liquidó al marketplace; el marketplace debe pagar al proveedor
    ("Venta",          "Ingreso"): "PAGAR_PROVEEDOR",
    ("Killer",         "Ingreso"): "PAGAR_PROVEEDOR",   # oferta financiada por Walmart

    # ── NOTA DE CRÉDITO ──────────────────────────────────────────────────────
    # Devolución del cliente; se revierte solo la venta
    ("Venta",          "Egreso"):  "NOTA_CREDITO",

    # ── GASTO DE PLATAFORMA ──────────────────────────────────────────────────
    # Costos operativos y comisiones deducibles
    ("Comision",                    "Egreso"): "GASTO_PLATAFORMA",
    ("killerCommission",            "Egreso"): "GASTO_PLATAFORMA",
    ("Killer",                      "Egreso"): "GASTO_PLATAFORMA",
    ("Campañas - Marketing SEM",    "Egreso"): "GASTO_PLATAFORMA",
    ("WFS - Tarifa de almacenaje",  "Egreso"): "GASTO_PLATAFORMA",
    ("WFS - Tarifa de envio",       "Egreso"): "GASTO_PLATAFORMA",
    ("WFS - Tarifa de manejo",      "Egreso"): "GASTO_PLATAFORMA",

    # ── RETENCIÓN FISCAL ─────────────────────────────────────────────────────
    # Egreso: Walmart retiene en ventas normales (no generan pago inmediato)
    # Ingreso: Walmart devuelve la retención por devolución del cliente
    ("Retencion IVA",  "Egreso"):  "RETENCION_FISCAL",
    ("Retencion ISR",  "Egreso"):  "RETENCION_FISCAL",
    ("Retencion IVA",  "Ingreso"): "RETENCION_FISCAL",
    ("Retencion ISR",  "Ingreso"): "RETENCION_FISCAL",
}

ETIQUETA_ACCION = {
    "PAGAR_PROVEEDOR":  "A pagar proveedor",
    "NOTA_CREDITO":     "Nota de crédito",
    "GASTO_PLATAFORMA": "Gasto de plataforma",
    "RETENCION_FISCAL": "Retención fiscal",
}

COLOR_ACCION = {
    "PAGAR_PROVEEDOR":  ("E2EFDA", "006100"),   # verde
    "NOTA_CREDITO":     ("FCE4D6", "843C0C"),   # naranja
    "GASTO_PLATAFORMA": ("FFF2CC", "7F6000"),   # amarillo
    "RETENCION_FISCAL": ("DEEAF1", "1F3864"),   # azul claro
    None:               ("F2F2F2", "595959"),   # gris (sin clasificar)
}


def clasificar_accion(concept: str, io: str) -> str | None:
    return ACCION_ERP.get((concept.strip(), io.strip()))


def enriquecer_df(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega columna Accion_ERP y ERP_ID/ERP_Tipo al DataFrame del CSV."""
    df = df.copy()
    df["Accion_ERP"] = df.apply(
        lambda r: clasificar_accion(r["Concept"], r["Income / Outcome"]), axis=1
    )
    df["Accion_Label"] = df["Accion_ERP"].map(
        lambda x: ETIQUETA_ACCION.get(x, "Sin clasificar")
    )
    return df


# ══════════════════════════════════════════════════════════════════════════════
#  ESTILOS COMPARTIDOS
# ══════════════════════════════════════════════════════════════════════════════

def hfill(h):  return PatternFill("solid", start_color=h, fgColor=h)
def thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)
def mfmt():    return '#,##0.00;(#,##0.00);"-"'

C = dict(
    blue="2E75B6", blue2="0070C0", lgray="F2F2F2", white="FFFFFF",
    yellow="FFF2CC", secbg="D6E4F0", green_bg="E2EFDA", red_bg="FFE0E0",
    blue_bg="DEEAF1", orange_bg="FCE4D6",
)
F = {
    "title": Font(name="Arial", bold=True, size=13, color="FFFFFF"),
    "hdr":   Font(name="Arial", bold=True, size=10, color="FFFFFF"),
    "body":  Font(name="Arial", size=10),
    "bold":  Font(name="Arial", bold=True, size=10),
    "sec":   Font(name="Arial", bold=True, size=10, color="1F3864"),
    "ital":  Font(name="Arial", size=9, italic=True, color="595959"),
    "grn":   Font(name="Arial", size=10, color="00B050"),
    "red":   Font(name="Arial", bold=True, size=10, color="FF0000"),
    "org":   Font(name="Arial", size=10, color="FF9900"),
}

COL_ANCHOS = {
    "Cuenta": 10,
    "ERP_ID": 26, "ERP_Tipo": 22, "Accion_ERP": 16, "Accion_Label": 22,
    "Payment Date": 14, "Orderline Number": 28, "Amount": 16,
    "Income / Outcome": 16, "Concept": 24, "FulfilledBy": 20,
    "Invoice Number": 18, "Return Number": 18, "Quantity": 10,
    "PartnerId": 14, "VendorNumber": 16, "Store": 30,
    "Date of Retention/Refund": 22, "Invoice Date": 14,
    "Credit Note": 14, "Date for CN": 14, "Fulfillment LineId": 20,
    "Shipment Number": 20, "Tracking Number": 28,
    "Total Walmart Funded Savings Program": 20,
}


def set_col_widths(ws, cols):
    for i, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = COL_ANCHOS.get(col, 16)


def banner(ws, texto, color, n_cols, row=1):
    ultima = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{ultima}{row}")
    ws[f"A{row}"] = texto
    ws[f"A{row}"].font = F["title"]
    ws[f"A{row}"].fill = hfill(color)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26


def escribir_encabezados(ws, cols, row=2):
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=col)
        c.font = F["hdr"]; c.fill = hfill(C["blue"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin()
    ws.row_dimensions[row].height = 26


def escribir_filas(ws, df, cols, start_row=3):
    for ri, row_data in enumerate(df[cols].itertuples(index=False), start_row):
        alt = hfill(C["lgray"]) if ri % 2 == 0 else hfill(C["white"])
        for ci, (col, val) in enumerate(zip(cols, row_data), 1):
            val = "" if (val is None or (isinstance(val, float) and pd.isna(val))) else val
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = thin()
            c.alignment = Alignment(vertical="center")

            # Color de fondo por acción
            if col == "Accion_Label":
                accion = df.iloc[ri - start_row]["Accion_ERP"] if "Accion_ERP" in df.columns else None
                bg, fg = COLOR_ACCION.get(accion, COLOR_ACCION[None])
                c.fill = hfill(bg)
                c.font = Font(name="Arial", size=10, bold=True, color=fg)
            elif col == "ERP_ID":
                if str(val).startswith("FULL-"):
                    c.fill = hfill(C["green_bg"])
                    c.font = Font(name="Arial", size=10, bold=True, color="006100")
                elif val and str(val) != "":
                    c.fill = hfill(C["blue_bg"])
                    c.font = Font(name="Arial", size=10, bold=True, color="0070C0")
                else:
                    c.fill = hfill(C["red_bg"])
                    c.font = Font(name="Arial", size=10, color="CC0000")
            elif col == "Amount":
                c.number_format = mfmt()
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.fill = alt
                c.font = F["body"]
            else:
                c.fill = alt
                c.font = F["body"]

        ws.row_dimensions[ri].height = 16


# ══════════════════════════════════════════════════════════════════════════════
#  CONSTRUCCIÓN DEL EXCEL UNIFICADO
# ══════════════════════════════════════════════════════════════════════════════

def hoja_resumen_ejecutivo(wb, df):
    """Hoja 1: Dashboard ejecutivo con resumen de acciones ERP."""
    ws = wb.create_sheet("📊 Resumen Ejecutivo")
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 32

    # Título
    ws.merge_cells("A1:E1")
    ws["A1"] = "REPORTE MENSUAL WALMART MX – ATLAS DEL DESCANSO"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = hfill(C["blue2"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Meta
    fecha = df["Payment Date"].dropna().iloc[0] if not df.empty else "N/D"
    for r, (lbl, val) in enumerate([
        ("Periodo de pago:", str(fecha)),
        ("Generado:", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Total registros CSV:", f"{len(df):,}"),
    ], 2):
        ws.cell(row=r, column=1, value=lbl).font = F["bold"]
        ws.cell(row=r, column=2, value=val).font  = F["body"]
        for ci in range(1, 6):
            ws.cell(row=r, column=ci).fill   = hfill(C["lgray"])
            ws.cell(row=r, column=ci).border = thin()
        ws.row_dimensions[r].height = 16

    # ── Resumen de Acciones ERP ────────────────────────────────────────────────
    r = 6
    ws.merge_cells(f"A{r}:E{r}")
    ws.cell(row=r, column=1, value="ACCIONES ERP – RESUMEN")
    ws.cell(row=r, column=1).font  = F["sec"]
    ws.cell(row=r, column=1).fill  = hfill(C["secbg"])
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(row=r, column=1).border = thin()
    ws.row_dimensions[r].height = 18; r += 1

    for ci, h in enumerate(["Acción", "Registros", "Monto Total ($)", "", ""], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = F["hdr"]; c.fill = hfill(C["blue"])
        c.alignment = Alignment(horizontal="center"); c.border = thin()
    ws.row_dimensions[r].height = 20; r += 1

    for accion, label in ETIQUETA_ACCION.items():
        mask    = df["Accion_ERP"] == accion
        cnt     = mask.sum()
        monto   = df.loc[mask, "Amount"].sum()
        bg, fg  = COLOR_ACCION[accion]
        for ci, val in enumerate([label, cnt, monto, "", ""], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = hfill(bg); c.border = thin()
            c.font = Font(name="Arial", size=10, bold=(ci == 1), color=fg)
            if ci == 3:
                c.number_format = mfmt(); c.alignment = Alignment(horizontal="right")
        ws.row_dimensions[r].height = 18; r += 1

    # Sin clasificar
    mask_sc = df["Accion_ERP"].isna()
    if mask_sc.any():
        for ci, val in enumerate(["Sin clasificar", mask_sc.sum(),
                                   df.loc[mask_sc, "Amount"].sum(), "", ""], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = hfill("F2F2F2"); c.border = thin()
            c.font = Font(name="Arial", size=10, color="595959")
            if ci == 3:
                c.number_format = mfmt(); c.alignment = Alignment(horizontal="right")
        ws.row_dimensions[r].height = 18; r += 1

    # Total ERP
    for ci, val in enumerate(["TOTAL", len(df), df["Amount"].sum(), "", ""], 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.fill = hfill(C["yellow"]); c.border = thin()
        c.font = F["bold"]
        if ci == 3:
            c.number_format = mfmt(); c.alignment = Alignment(horizontal="right")
    ws.row_dimensions[r].height = 18; r += 2

    # ── Bloque: Desglose por Concepto / Tipo ──────────────────────────────────
    ws.merge_cells(f"A{r}:E{r}")
    ws.cell(row=r, column=1, value="DESGLOSE POR CONCEPTO – CSV vs ERP")
    ws.cell(row=r, column=1).font      = F["sec"]
    ws.cell(row=r, column=1).fill      = hfill(C["secbg"])
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(row=r, column=1).border    = thin()
    ws.row_dimensions[r].height = 18; r += 1

    for ci, h in enumerate(["Concepto", "Tipo", "Total CSV ($)", "Registros", ""], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = F["hdr"]; c.fill = hfill(C["blue"])
        c.alignment = Alignment(horizontal="center"); c.border = thin()
    ws.row_dimensions[r].height = 20; r += 1

    combos = (
        df.groupby(["Concept", "Income / Outcome"])["Amount"]
        .agg(["sum", "count"])
        .reset_index()
        .sort_values(["Income / Outcome", "Concept"], ascending=[False, True])
    )
    neto_total = 0.0
    for _, row_data in combos.iterrows():
        concepto = row_data["Concept"]
        tipo     = row_data["Income / Outcome"]
        total    = row_data["sum"]
        cnt      = int(row_data["count"])
        neto_total += total
        fill = hfill(C["lgray"]) if r % 2 == 0 else hfill(C["white"])
        fg   = "006100" if tipo == "Ingreso" else "C00000"
        for ci, val in enumerate([concepto, tipo, total, cnt, ""], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = fill; c.border = thin()
            c.font = Font(name="Arial", size=10, color=fg if ci <= 2 else "000000")
            if ci == 3:
                c.number_format = mfmt(); c.alignment = Alignment(horizontal="right")
            if ci == 4:
                c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 16; r += 1

    # Neto CSV total
    for ci, val in enumerate(["Neto Total CSV", "", neto_total, "", ""], 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.fill = hfill(C["yellow"]); c.border = thin()
        c.font = F["bold"]
        if ci == 3:
            c.number_format = mfmt(); c.alignment = Alignment(horizontal="right")
    ws.row_dimensions[r].height = 18; r += 1

    # ERP total vs Venta Ingreso
    if "ERP_PrecioTotal" in df.columns:
        df_tmp = df.copy()
        df_tmp["_oid"] = df_tmp["Orderline Number"].apply(_erp.extraer_order_id)
        erp_total = pd.to_numeric(
            df_tmp.groupby("_oid")["ERP_PrecioTotal"].first(), errors="coerce"
        ).sum()
        mask_v = (df["Concept"] == "Venta") & (df["Income / Outcome"] == "Ingreso")
        venta_total = df.loc[mask_v, "Amount"].sum()
        diff_erp = round(venta_total - float(erp_total), 2)
        ok = "✅ Cuadra" if abs(diff_erp) < 1 else f"❌ Dif ${diff_erp:+,.2f}"

        for ci, val in enumerate(["ERP Total (facturas/pedidos)", "", erp_total, "", ""], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = hfill(C["blue_bg"]); c.border = thin(); c.font = F["bold"]
            if ci == 3:
                c.number_format = mfmt(); c.alignment = Alignment(horizontal="right")
        ws.row_dimensions[r].height = 16; r += 1

        for ci, val in enumerate(["  Venta Ingreso CSV vs ERP", "", diff_erp, "", ok], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = hfill(C["green_bg"] if abs(diff_erp) < 1 else C["red_bg"])
            c.border = thin()
            c.font = Font(name="Arial", size=10,
                          color="006100" if abs(diff_erp) < 1 else "C00000")
            if ci == 3:
                c.number_format = mfmt(); c.alignment = Alignment(horizontal="right")
            if ci == 5:
                c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 16


def hoja_accion(wb, df, accion_key, titulo_hoja, titulo_banner, color_banner):
    """Genera una hoja filtrada por acción ERP."""
    df_f = df[df["Accion_ERP"] == accion_key].copy()
    if df_f.empty:
        return

    ws = wb.create_sheet(titulo_hoja)

    cols_prio  = ["ERP_ID", "Accion_Label", "Payment Date", "Orderline Number",
                  "Amount", "Concept", "Income / Outcome", "FulfilledBy",
                  "Invoice Number", "Return Number", "Quantity"]
    if "Cuenta" in df_f.columns:
        cols_prio = ["Cuenta"] + cols_prio
    cols_extra = [c for c in df_f.columns
                  if c not in cols_prio and c not in ("Accion_ERP",)]
    cols = [c for c in cols_prio if c in df_f.columns] + cols_extra

    set_col_widths(ws, cols)
    banner(ws, titulo_banner, color_banner, len(cols), row=1)

    # Info
    ws.cell(row=2, column=1, value="Registros:").font = F["bold"]
    ws.cell(row=2, column=2, value=len(df_f)).font    = F["body"]
    ws.cell(row=2, column=3, value="Monto total:").font = F["bold"]
    ws.cell(row=2, column=4, value=df_f["Amount"].sum()).font = F["body"]
    ws.cell(row=2, column=4).number_format = mfmt()
    for ci in range(1, len(cols) + 1):
        ws.cell(row=2, column=ci).fill   = hfill(C["lgray"])
        ws.cell(row=2, column=ci).border = thin()
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    escribir_encabezados(ws, cols, row=4)
    escribir_filas(ws, df_f, cols, start_row=5)


def hoja_pendientes(wb, df):
    """Hoja de órdenes Seller sin PO# (seguimiento manual)."""
    mask = (
        df["FulfilledBy"].str.lower().str.contains("seller", na=False) &
        (df["ERP_ID"].isna() | df["ERP_ID"].eq(""))
    )
    df_p = df[mask].copy()
    if df_p.empty:
        return

    ws = wb.create_sheet("⚠ ERP_Pendientes")
    cols_prio  = ["Accion_Label", "Payment Date", "Orderline Number", "Amount",
                  "Concept", "Income / Outcome", "FulfilledBy",
                  "Invoice Number", "Return Number"]
    if "Cuenta" in df_p.columns:
        cols_prio = ["Cuenta"] + cols_prio
    cols_extra = [c for c in df_p.columns
                  if c not in cols_prio and c not in ("Accion_ERP", "ERP_ID", "ERP_Tipo")]
    cols = [c for c in cols_prio if c in df_p.columns] + cols_extra

    set_col_widths(ws, cols)
    banner(ws, "⚠  PENDIENTES – Seller Fulfillment sin PO# (seguimiento manual)",
           "C00000", len(cols), row=1)
    ws.cell(row=2, column=1, value="Registros:").font  = F["bold"]
    ws.cell(row=2, column=2, value=len(df_p)).font     = F["body"]
    ws.cell(row=2, column=3, value="Monto:").font      = F["bold"]
    ws.cell(row=2, column=4, value=df_p["Amount"].sum()).font = F["body"]
    ws.cell(row=2, column=4).number_format = mfmt()
    for ci in range(1, len(cols) + 1):
        ws.cell(row=2, column=ci).fill   = hfill(C["lgray"])
        ws.cell(row=2, column=ci).border = thin()
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    escribir_encabezados(ws, cols, row=4)
    escribir_filas(ws, df_p, cols, start_row=5)


def hoja_detalle_completo(wb, df):
    """Hoja con todos los registros enriquecidos (ERP_ID + Accion_Label)."""
    ws = wb.create_sheet("📋 Detalle Completo")

    cols_prio  = ["ERP_ID", "ERP_Tipo", "Accion_Label",
                  "Payment Date", "Orderline Number", "Amount",
                  "Income / Outcome", "Concept", "FulfilledBy",
                  "Invoice Number", "Return Number", "Quantity"]
    cols_extra = [c for c in df.columns
                  if c not in cols_prio and c not in ("Accion_ERP",)]
    cols = [c for c in cols_prio if c in df.columns] + cols_extra

    set_col_widths(ws, cols)
    banner(ws, "DETALLE COMPLETO – Todos los registros con ERP_ID y clasificación contable",
           C["blue2"], len(cols), row=1)
    ws.row_dimensions[2].height = 6
    escribir_encabezados(ws, cols, row=3)
    escribir_filas(ws, df, cols, start_row=4)


def crear_excel_unificado(df, sin_mapeo, output: str):
    wb = Workbook()
    del wb["Sheet"]  # elimina hoja vacía default

    # 1. Resumen ejecutivo (resumen de acciones ERP)
    hoja_resumen_ejecutivo(wb, df)

    # 2. Hojas por acción contable
    hoja_accion(wb, df,
                "PAGAR_PROVEEDOR",
                "💰 A Pagar Proveedor",
                "A PAGAR PROVEEDOR – Walmart liquidó estas órdenes, enviar a pago",
                "375623")

    hoja_accion(wb, df,
                "NOTA_CREDITO",
                "🔄 Notas de Crédito",
                "NOTAS DE CRÉDITO – Devoluciones recibidas de Walmart",
                "843C0C")

    hoja_accion(wb, df,
                "GASTO_PLATAFORMA",
                "📉 Gastos Plataforma",
                "GASTOS DE PLATAFORMA – Comisiones, SEM, WFS, Killer",
                "7F6000")

    hoja_accion(wb, df,
                "RETENCION_FISCAL",
                "🏛 Retenciones Fiscales",
                "RETENCIONES FISCALES – IVA e ISR retenidos/devueltos (Egreso: retención | Ingreso: devolución por reembolso)",
                "1F3864")

    # 3. Pendientes sin PO#
    hoja_pendientes(wb, df)

    # 4. Detalle completo enriquecido
    hoja_detalle_completo(wb, df)

    # 5. Conciliación ERP (si se consultó el ERP)
    if "ERP_En_Sistema" in df.columns:
        _erp._hoja_erp_conciliacion(wb, df)

    # 6. Hoja sin mapeo (auditoría)
    if sin_mapeo:
        ws = wb.create_sheet("🔍 Sin Mapeo")
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        for ci, h in enumerate(["Concepto CSV", "Tipo", "Total"], 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = F["hdr"]; c.fill = hfill(C["blue"])
            c.alignment = Alignment(horizontal="center"); c.border = thin()
        for ri, ((concepto, tipo), total) in enumerate(sin_mapeo.items(), 2):
            alt = hfill(C["lgray"]) if ri % 2 == 0 else hfill(C["white"])
            for ci, val in enumerate([concepto, tipo, round(total, 2)], 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = F["body"]; c.fill = alt; c.border = thin()
                if ci == 3:
                    c.number_format = mfmt()
                    c.alignment = Alignment(horizontal="right")

    wb.save(output)


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Orquestador mensual de conciliación Walmart MX"
    )
    parser.add_argument("csv",         help="Ruta al archivo CSV de pagos Walmart")
    parser.add_argument("--output", "-o", default="",
                        help="Nombre del Excel de salida")
    parser.add_argument("--skip-erp",  action="store_true",
                        help="Omite la consulta al ERP por completo")
    parser.add_argument("--cuenta", default="atlas",
                        help="Cuenta Walmart a procesar: atlas (default) | spring")
    args = parser.parse_args()

    if not Path(args.csv).exists():
        print(f"\n  ERROR: No se encontró: {args.csv}"); sys.exit(1)

    output = args.output or Path(args.csv).stem + "_reporte.xlsx"

    print()
    print("  ╔══════════════════════════════════════════════════════════════╗")
    print("  ║      PROCESO MENSUAL CONCILIACIÓN WALMART MX                 ║")
    print("  ╚══════════════════════════════════════════════════════════════╝")

    # ── PASO 1: Leer CSV + enriquecer ─────────────────────────────────────────
    print("\n  ── PASO 1/3: Procesando CSV ────────────────────────────────────")
    df = _conc.leer_csv(args.csv)
    print(f"  → {len(df):,} registros cargados")

    _, sin_mapeo = _conc.agrupar_por_categoria(df)
    df = _erp.preparar_para_erp(df, cuenta=args.cuenta)
    df = enriquecer_df(df)

    # ── PASO 2: Consulta ERP ──────────────────────────────────────────────────
    if not args.skip_erp:
        print("\n  ── PASO 2/3: Consultando ERP ───────────────────────────────────")
        df = _erp.enriquecer_con_erp(df, cuenta=args.cuenta)
    else:
        print("\n  ── PASO 2/3: ERP omitido (--skip-erp) ──────────────────────────")

    # Resumen acciones ERP en consola
    print()
    print("  " + "═" * 54)
    print(f"  {'ACCIONES ERP':^54}")
    print("  " + "═" * 54)
    for accion, label in ETIQUETA_ACCION.items():
        mask  = df["Accion_ERP"] == accion
        cnt   = mask.sum()
        monto = df.loc[mask, "Amount"].sum()
        print(f"  {'▸'} {label:<28} {cnt:>5,} registros  ${monto:>14,.2f}")
    mask_sc = df["Accion_ERP"].isna()
    if mask_sc.any():
        print(f"  {'▸'} {'Sin clasificar':<28} {mask_sc.sum():>5,} registros  "
              f"${df.loc[mask_sc, 'Amount'].sum():>14,.2f}")

    # Pendientes sin PO#
    pend = (df["FulfilledBy"].str.lower().str.contains("seller", na=False) &
            (df["ERP_ID"].isna() | df["ERP_ID"].eq(""))).sum()
    if pend:
        print(f"\n  ⚠  {pend} registros Seller sin PO# → hoja 'ERP_Pendientes'")
    print("  " + "═" * 54)

    # ── PASO 3: Generar Excel unificado ───────────────────────────────────────
    print(f"\n  ── PASO 3/3: Generando Excel ───────────────────────────────────")
    print(f"  → {output}")
    crear_excel_unificado(df, sin_mapeo, output)
    print(f"  ✅ Listo: {output}")
    print()
    print("  Hojas generadas:")
    print("    📊 Resumen Ejecutivo      – Conciliación + resumen de acciones")
    print("    💰 A Pagar Proveedor      – Órdenes liquidadas por Walmart")
    print("    🔄 Notas de Crédito       – Devoluciones a registrar")
    print("    📉 Gastos Plataforma      – Comisiones, SEM, WFS, Killer")
    print("    🏛 Retenciones Fiscales   – IVA e ISR retenidos/devueltos (ver columna Income/Outcome)")
    if pend:
        print("    ⚠  ERP_Pendientes         – Seller sin PO# (seguimiento manual)")
    print("    📋 Detalle Completo       – Todos los registros enriquecidos")
    if "ERP_En_Sistema" in df.columns:
        print("    🔗 ERP_Conciliacion       – Resumen por orden CSV vs ERP (✅/⚠/❌)")
    if sin_mapeo:
        print("    🔍 Sin Mapeo              – Movimientos no clasificados")
    print()


if __name__ == "__main__":
    main()
